# -*- coding: utf-8 -*-
"""
status_tool.py —— 初筛/研判状态读写工具（增量机制核心，融合新增）

状态载体双轨（方案 4.3）：
1. 采集主文件（{地市}政府采购公告.xlsx）：第 11 列「初筛状态」、第 12 列「研判状态」；
   多 Sheet 主文件（政府采购公告.xlsx，不同地区不同 Sheet）时遍历全部 Sheet（含「初筛状态」列的 Sheet）；
2. 用户自备清单：清单路径/增量状态.json（用户清单文件本身只读，绝不改写）。

唯一键：地址链接URL（第 10 列）。回写一律按 URL 定位行，禁止按行号
（merge 排序会变动行号，按行号回写会错位）。

对外接口（方案 4.3 示意，本文件为实现）：
- read_state(list_dir, main_file=None) -> (state, warnings)
    state = {url: {"初筛状态": str, "研判状态": str}}；warnings = [中文提示...]
- write_state(list_dir, updates, main_file=None)
    updates = [{"url": ..., "初筛状态": ..., "研判状态": ...}]（字段可省略=不回写）
    目标 = 清单路径下状态载体：优先主文件（含"初筛状态"列），否则 增量状态.json
- reset_status(list_dir, main_file=None)  复评：全部记录重置为「新增待初筛 / 空」

容错（方案 9，全部保留）：
- 主文件读写一律 openpyxl.load_workbook(path) 不带 read_only=True（Python 3.13 崩溃坑位）；
- 写前检查锁文件 ~${文件名}，被 Excel 占用时中文提示并中止本次回写；
- 状态 JSON 损坏 → 视为无状态（全部=新增），中文提示；
- 清单缺 URL 列 → 无法按 URL 建状态（返回无状态，由 run_pipeline 降级全量并提示）。
"""

import json
import os

import openpyxl

# 状态取值（方案 4.1/4.2）
STATUS_SCREENED = "已初筛"
STATUS_NEW = "新增待初筛"
STATUS_JUDGED = "已研判"
STATUS_SKIPPED = "已跳过"

# 用户自备清单的状态载体文件名（写在清单路径下，方案 4.3）
STATUS_JSON_NAME = "增量状态.json"

# 主文件固定结构：第 1 行标题、第 2 行汇总行、第 3 行表头、第 4 行起数据
MAIN_HEADER_ROW = 3
MAIN_FIRST_DATA_ROW = 4

# 主文件数据列列名（10 数据列 + 2 状态列，方案 4.1）
MAIN_DATA_COLUMNS = [
    "公告时间", "公告类型", "所在区县", "采购人", "采购项目名称",
    "预算金额(元)", "中标金额(元)", "中标单位", "代理机构", "地址链接URL",
]
URL_COLUMN_NAME = "地址链接URL"
STATUS_SCREENED_COL = "初筛状态"
STATUS_JUDGED_COL = "研判状态"


class StatusToolError(Exception):
    """状态读写过程中的中文可读异常。"""


def _scan_main_file(list_dir: str):
    """
    扫描清单路径，返回含「初筛状态」列的主文件绝对路径；找不到返回 None。
    排除：锁文件（~$）、初筛结果/研判结果/高亮版、增量状态.json（非 xlsx）。
    """
    if not os.path.isdir(list_dir):
        return None
    for fname in sorted(os.listdir(list_dir)):
        if not fname.lower().endswith(".xlsx"):
            continue
        if fname.startswith("~$"):
            continue
        if "初筛结果" in fname or "研判结果" in fname or "高亮版" in fname:
            continue
        path = os.path.join(list_dir, fname)
        if _has_status_column(path):
            return path
    return None


def _sheet_headers(path: str):
    """
    遍历主文件全部 Sheet，返回 [{sheet: Sheet名, header_row: 表头行号, col_of: {列名: 列号}}]。
    表头行探测：前 5 行内找到含「地址链接URL」或「采购项目名称」的行（取先出现者）；
    不含数据的 Sheet（无 URL 列）跳过。失败返回 []。
    """
    sheets = []
    try:
        wb = openpyxl.load_workbook(path)  # read_only=True removed
        for ws in wb.worksheets:
            header_row = None
            for r in range(1, min(ws.max_row, 5) + 1):
                vals = [str(ws.cell(r, c).value or "").strip() for c in range(1, ws.max_column + 1)]
                if URL_COLUMN_NAME in vals or "采购项目名称" in vals:
                    header_row = r
                    break
            if header_row is None:
                continue
            col_of = {}
            for c in range(1, ws.max_column + 1):
                name = str(ws.cell(header_row, c).value or "").strip()
                if name:
                    col_of.setdefault(name, c)
            if URL_COLUMN_NAME not in col_of:
                continue
            sheets.append({"sheet": ws.title, "ws": ws, "header_row": header_row, "col_of": col_of})
        wb.close()
    except Exception:
        return []
    return sheets


def _has_status_column(path: str) -> bool:
    """任一 Sheet 前 5 行内是否含「初筛状态」列（带容错，失败视为无状态列）。"""
    try:
        wb = openpyxl.load_workbook(path)  # read_only=True removed（Python 3.13 崩溃坑位）
        for ws in wb.worksheets:
            for r in range(1, min(ws.max_row, 5) + 1):
                for c in range(1, ws.max_column + 1):
                    if str(ws.cell(r, c).value or "").strip() == STATUS_SCREENED_COL:
                        wb.close()
                        return True
        wb.close()
    except Exception:
        return False
    return False


def _lock_file_of(main_path: str) -> str:
    """返回主文件对应的 Excel 锁文件路径（~$文件名）。"""
    directory, name = os.path.split(main_path)
    return os.path.join(directory, "~$" + name)


def _check_lock(main_path: str) -> None:
    """写前检查锁文件；被 Excel 占用时中文提示并中止本次回写。"""
    lock = _lock_file_of(main_path)
    if os.path.exists(lock):
        raise StatusToolError(
            "错误：主文件被 Excel 打开，无法写入状态，请先关闭 Excel 再运行。\n"
            "  主文件：%s\n  锁文件：%s\n"
            "（本次状态未回写，不影响已完成的初筛/研判产物；关闭 Excel 后重跑一次初筛即可补齐）"
            % (main_path, lock)
        )


def _load_status_json(list_dir: str, warnings: list) -> dict:
    """读取 清单路径/增量状态.json；不存在返回 {}；损坏时中文提示并视为无状态。"""
    path = os.path.join(list_dir, STATUS_JSON_NAME)
    if not os.path.exists(path):
        return {}
    try:
        with open(path, encoding="utf-8-sig") as f:
            data = json.load(f)
    except Exception:
        warnings.append(
            "增量状态文件无法读取（%s），本次按全部新增处理；如文件损坏可删除后由技能重建。"
            % path
        )
        return {}
    if not isinstance(data, dict):
        warnings.append(
            "增量状态文件格式无法识别（%s），本次按全部新增处理。" % path
        )
        return {}
    records = data.get("records")
    if not isinstance(records, dict):
        return {}
    state = {}
    for url, st in records.items():
        if not isinstance(st, dict):
            continue
        state[str(url).strip()] = {
            STATUS_SCREENED_COL: str(st.get(STATUS_SCREENED_COL, "") or ""),
            STATUS_JUDGED_COL: str(st.get(STATUS_JUDGED_COL, "") or ""),
        }
    return state


def _load_main_state(path: str, warnings: list) -> dict:
    """从主文件列读取状态（遍历全部含 URL 列的 Sheet）：{url: {初筛状态, 研判状态}}。"""
    sheets = _sheet_headers(path)
    if not sheets:
        warnings.append(
            "主文件「%s」缺少地址链接URL列，无法按 URL 增量，本次按全部新增处理。" % os.path.basename(path)
        )
        return {}
    state = {}
    try:
        wb = openpyxl.load_workbook(path)  # read_only=True removed
        for ws in wb.worksheets:
            header_row = None
            for r in range(1, min(ws.max_row, 5) + 1):
                vals = [str(ws.cell(r, c).value or "").strip() for c in range(1, ws.max_column + 1)]
                if URL_COLUMN_NAME in vals:
                    header_row = r
                    break
            if header_row is None:
                continue
            col_of = {}
            for c in range(1, ws.max_column + 1):
                name = str(ws.cell(header_row, c).value or "").strip()
                if name:
                    col_of.setdefault(name, c)
            url_col = col_of.get(URL_COLUMN_NAME)
            if not url_col:
                continue
            scr_col = col_of.get(STATUS_SCREENED_COL)
            jdg_col = col_of.get(STATUS_JUDGED_COL)
            for r in range(header_row + 1, ws.max_row + 1):
                url = str(ws.cell(r, url_col).value or "").strip()
                if not url:
                    continue
                state[url] = {
                    STATUS_SCREENED_COL: str(ws.cell(r, scr_col).value or "").strip() if scr_col else "",
                    STATUS_JUDGED_COL: str(ws.cell(r, jdg_col).value or "").strip() if jdg_col else "",
                }
        wb.close()
    except Exception as e:
        warnings.append("读取主文件状态失败（%s）：%s，本次按全部新增处理。" % (path, e))
        return {}
    return state


def read_state(list_dir: str, main_file: str = None):
    """
    读取清单路径下的状态载体（主文件列 + 增量状态.json 合并，按 URL 唯一键）。

    返回 (state, warnings)：
    - state = {url: {"初筛状态": str, "研判状态": str}}
    - warnings = [中文提示...]（损坏/缺 URL 列等容错提示）
    """
    warnings = []
    state = {}
    target = main_file if main_file and os.path.exists(main_file) else _scan_main_file(list_dir)
    if target:
        state.update(_load_main_state(target, warnings))
    state.update(_load_status_json(list_dir, warnings))
    return state, warnings


def write_state(list_dir: str, updates: list, main_file: str = None) -> None:
    """
    回写状态到清单路径下的状态载体：
    优先主文件（含「初筛状态」列）；否则写 清单路径/增量状态.json。
    updates = [{"url": ..., "初筛状态": ..., "研判状态": ...}]，字段省略=不回写该列。
    按 URL 定位行，禁止按行号；写前检查锁文件；读写不带 read_only=True。
    """
    valid = [
        u for u in updates
        if isinstance(u, dict) and str(u.get("url", "") or "").strip()
    ]
    if not valid:
        return
    target = main_file if main_file and os.path.exists(main_file) else _scan_main_file(list_dir)
    if target:
        _write_main_state(target, valid)
        return
    _write_status_json(list_dir, valid)


def _write_main_state(main_path: str, updates: list) -> None:
    """回写主文件状态列（遍历全部含 URL 列的 Sheet，按 URL 定位行）。"""
    _check_lock(main_path)
    sheets = _sheet_headers(main_path)
    if not sheets:
        raise StatusToolError(
            "主文件「%s」缺少地址链接URL列，无法回写状态；已改按 清单路径/%s 记录（可手工编辑查看）。"
            % (os.path.basename(main_path), STATUS_JSON_NAME)
        )

    # 预建 URL → [(Sheet, 列名→列号, 行号)...] 映射（本次回写内使用；行号不落盘，仅用于本次定位）。
    # 同一 URL 可能出现在多个 Sheet（市本级/区县重复挂载）：更新时写入全部命中的 Sheet，
    # 避免仅按 URL 建映射被后遍历的 Sheet 覆盖而回写错位。
    url_to_locs = {}
    try:
        wb = openpyxl.load_workbook(main_path)  # read_only=True removed
        for ws in wb.worksheets:
            header_row = None
            for r in range(1, min(ws.max_row, 5) + 1):
                vals = [str(ws.cell(r, c).value or "").strip() for c in range(1, ws.max_column + 1)]
                if URL_COLUMN_NAME in vals:
                    header_row = r
                    break
            if header_row is None:
                continue
            col_of = {}
            for c in range(1, ws.max_column + 1):
                name = str(ws.cell(header_row, c).value or "").strip()
                if name:
                    col_of.setdefault(name, c)
            url_col = col_of.get(URL_COLUMN_NAME)
            if not url_col:
                continue
            for r in range(header_row + 1, ws.max_row + 1):
                url = str(ws.cell(r, url_col).value or "").strip()
                if url:
                    url_to_locs.setdefault(url, []).append((ws, col_of, r))
        changed = 0
        for u in updates:
            url = str(u.get("url", "") or "").strip()
            locs = url_to_locs.get(url)
            if not locs:
                continue  # 该 URL 已不在主文件（理论不出现），跳过
            for ws, col_of, row in locs:
                scr_col = col_of.get(STATUS_SCREENED_COL)
                jdg_col = col_of.get(STATUS_JUDGED_COL)
                if scr_col is not None and u.get(STATUS_SCREENED_COL) is not None:
                    ws.cell(row, scr_col).value = u[STATUS_SCREENED_COL]
                    changed += 1
                if jdg_col is not None and u.get(STATUS_JUDGED_COL) is not None:
                    ws.cell(row, jdg_col).value = u[STATUS_JUDGED_COL]
                    changed += 1
        # 保存前再次检查锁：写入循环耗时较长，期间文件可能被 Excel 打开（锁在循环中才出现）
        _check_lock(main_path)
        wb.save(main_path)
        wb.close()
    except StatusToolError:
        raise
    except PermissionError:
        raise StatusToolError(
            "保存主文件失败：%s 可能已被 Excel 打开占用（或被杀毒软件/云同步临时锁定），"
            "请先关闭 Excel 再运行；若已关闭仍报此错，请稍等几秒后重试。" % main_path
        )
    except Exception as e:
        raise StatusToolError("回写主文件状态失败（%s）：%s" % (main_path, e))


def _write_status_json(list_dir: str, updates: list) -> None:
    """写入/合并 清单路径/增量状态.json（用户清单只读，状态独立落盘）。"""
    warnings = []
    records = _load_status_json(list_dir, warnings)  # 损坏时按空重建
    for u in updates:
        url = str(u.get("url", "") or "").strip()
        if not url:
            continue
        st = records.setdefault(url, {STATUS_SCREENED_COL: "", STATUS_JUDGED_COL: ""})
        if u.get(STATUS_SCREENED_COL) is not None:
            st[STATUS_SCREENED_COL] = u[STATUS_SCREENED_COL]
        if u.get(STATUS_JUDGED_COL) is not None:
            st[STATUS_JUDGED_COL] = u[STATUS_JUDGED_COL]
    data = {
        "说明": "用户自备清单的初筛/研判状态（按地址链接URL记录）。由技能自动维护，用户无需编辑。",
        "records": records,
    }
    path = os.path.join(list_dir, STATUS_JSON_NAME)
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception as e:
        raise StatusToolError("保存增量状态文件失败（%s）：%s" % (path, e))


def reset_status(list_dir: str, main_file: str = None) -> None:
    """
    复评：将状态载体中全部记录重置为「初筛状态=新增待初筛、研判状态=空」，
    下一轮按全量新增处理。主文件列与 增量状态.json 同时重置。
    """
    target = main_file if main_file and os.path.exists(main_file) else _scan_main_file(list_dir)
    if target:
        _reset_main_status(target)
    json_path = os.path.join(list_dir, STATUS_JSON_NAME)
    if os.path.exists(json_path):
        try:
            data = {
                "说明": "用户自备清单的初筛/研判状态（按地址链接URL记录）。由技能自动维护，用户无需编辑。",
                "records": {},
            }
            with open(json_path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            raise StatusToolError("重置增量状态文件失败（%s）：%s" % (json_path, e))


def _reset_main_status(main_path: str) -> None:
    """重置主文件全部数据行的状态列（遍历全部含 URL 列的 Sheet）。"""
    _check_lock(main_path)
    sheets = _sheet_headers(main_path)
    if not sheets:
        raise StatusToolError(
            "主文件「%s」缺少地址链接URL列，无法重置状态。" % os.path.basename(main_path)
        )
    try:
        wb = openpyxl.load_workbook(main_path)  # read_only=True removed
        reset = 0
        for ws in wb.worksheets:
            header_row = None
            for r in range(1, min(ws.max_row, 5) + 1):
                vals = [str(ws.cell(r, c).value or "").strip() for c in range(1, ws.max_column + 1)]
                if URL_COLUMN_NAME in vals:
                    header_row = r
                    break
            if header_row is None:
                continue
            col_of = {}
            for c in range(1, ws.max_column + 1):
                name = str(ws.cell(header_row, c).value or "").strip()
                if name:
                    col_of.setdefault(name, c)
            url_col = col_of.get(URL_COLUMN_NAME)
            if not url_col:
                continue
            scr_col = col_of.get(STATUS_SCREENED_COL)
            jdg_col = col_of.get(STATUS_JUDGED_COL)
            for r in range(header_row + 1, ws.max_row + 1):
                url = str(ws.cell(r, url_col).value or "").strip()
                if not url:
                    continue
                if scr_col is not None:
                    ws.cell(r, scr_col).value = STATUS_NEW
                if jdg_col is not None:
                    ws.cell(r, jdg_col).value = ""
                reset += 1
        # 保存前再次检查锁：写入循环耗时较长，期间文件可能被 Excel 打开（锁在循环中才出现）
        _check_lock(main_path)
        wb.save(main_path)
        wb.close()
    except StatusToolError:
        raise
    except PermissionError:
        raise StatusToolError(
            "保存主文件失败：%s 可能已被 Excel 打开占用（或被杀毒软件/云同步临时锁定），"
            "请先关闭 Excel 再运行；若已关闭仍报此错，请稍等几秒后重试。" % main_path
        )
    except Exception as e:
        raise StatusToolError("重置主文件状态失败（%s）：%s" % (main_path, e))
