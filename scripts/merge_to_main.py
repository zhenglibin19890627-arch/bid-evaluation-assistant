# -*- coding: utf-8 -*-
"""
合并采集结果到主文件
将所有 {地市}_采购公告_*.xlsx 合并到 {地市}政府采购公告.xlsx
或（--main-file 指向多 Sheet 主文件时）合并到「政府采购公告.xlsx」对应地市 Sheet
（用户单文件多 Sheet 工作流：不同地区不同 Sheet，初筛结果同表回写，不新建表格）。

融合改造（方案 7.2/8）：
- --city 参数（默认"丽水市"），主文件/lock/批次文件均按地市命名；
- 新增行写入状态列：初筛状态=新增待初筛、研判状态=空；
- 排序重写搬运 1–12 列（含状态列），状态列随行走，避免与重排后数据错位；
- 首次运行主文件不存在时自动创建空主文件（第 1 行标题、第 2 行汇总行占位、
  第 3 行表头 12 列、第 4 行起数据），零代码首次运行不中断。

多 Sheet 模式（--main-file）：
- 主文件为「政府采购公告.xlsx」，Sheet 名 = 地市名（如"丽水"）；
- 每个 Sheet 统一 15 列：10 数据列 + 初筛状态/研判状态 + 初筛结果/初筛原因/相关性(命中词)，
  初筛结果直接写回主文件对应 Sheet（不再创建 初筛结果.xlsx 新表格）；
- 旧 Sheet 列数不足时自动补齐列（缺列置空）。

用法：
  python merge_to_main.py --workspace C:\\path\\to\\workspace
  python merge_to_main.py --workspace .
  python merge_to_main.py --city 杭州市 --workspace C:\\path\\to\\workspace
  python merge_to_main.py --city 丽水市 --main-file "C:\\path\\to\\workspace\\政府采购公告.xlsx" --workspace C:\\path\\to\\workspace
"""
import openpyxl
import os
import glob
import sys
import argparse
from datetime import datetime
from pathlib import Path
from openpyxl.styles import Font, Alignment

# 主文件结构：第 1 行标题、第 2 行汇总行、第 3 行表头、第 4 行起数据
HEADER_ROW = 3
FIRST_DATA_ROW = 4

# 主文件 12 列：10 数据列 + 初筛状态/研判状态（方案 4.1）
DATA_COLUMNS = [
    '公告时间', '公告类型', '所在区县', '采购人', '采购项目名称',
    '预算金额(元)', '中标金额(元)', '中标单位', '代理机构', '地址链接URL',
]
STATUS_SCREENED_COL = '初筛状态'
STATUS_JUDGED_COL = '研判状态'
ALL_COLUMNS = DATA_COLUMNS + [STATUS_SCREENED_COL, STATUS_JUDGED_COL]

# 多 Sheet 主文件（政府采购公告.xlsx）15 列：10 数据列 + 2 状态列 + 初筛结果 3 列
SCREEN_RESULT_COL = '初筛结果'
SCREEN_REASON_COL = '初筛原因'
RELEVANCE_COL = '相关性(命中词)'
MULTI_MAIN_COLUMNS = ALL_COLUMNS + [SCREEN_RESULT_COL, SCREEN_REASON_COL, RELEVANCE_COL]

# 多 Sheet 主文件名（用户单文件工作流：不同地区不同 Sheet）
MULTI_MAIN_FILE_NAME = '政府采购公告.xlsx'

COL_WIDTHS = {'A': 16, 'B': 14, 'C': 14, 'D': 25, 'E': 45,
              'F': 14, 'G': 14, 'H': 25, 'I': 25, 'J': 50, 'K': 12, 'L': 12}
MULTI_COL_WIDTHS = dict(COL_WIDTHS)
MULTI_COL_WIDTHS.update({'M': 12, 'N': 30, 'O': 30})


def get_args():
    parser = argparse.ArgumentParser(
        description='合并批次文件到主文件',
        epilog='Example: python merge_to_main.py --workspace C:\\path\\to\\workspace'
    )
    parser.add_argument('--workspace', type=str, default='.',
                        help='工作目录（主文件 + 批次文件所在），默认当前目录 .')
    parser.add_argument('--city', type=str, default='丽水市',
                        help='目标地市（默认丽水市），主文件/批次文件按地市命名')
    parser.add_argument('--main-file', type=str, default=None,
                        help='多 Sheet 主文件完整路径（如 …\\政府采购公告.xlsx）；'
                             '指定时合并到该文件对应地市 Sheet（不同地区不同 Sheet，初筛同表回写）')
    return parser.parse_args()


def _col_index(ws, header_row):
    """按表头行列名查找列号（工程约定：列名常量替代魔法列号）。返回 {列名: 列号}。"""
    col_of = {}
    for c in range(1, ws.max_column + 1):
        name = str(ws.cell(header_row, c).value or '').strip()
        if name:
            col_of.setdefault(name, c)
    return col_of


def _normalize_time(val):
    """时间值 → 文本 'YYYY-MM-DD HH:MM'（不带秒）。"""
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d %H:%M')
    if isinstance(val, str) and len(val) >= 16:
        return val[:16]
    return val


def _sort_key(val):
    """稳健排序键：兼容 datetime 与 'YYYY-MM-DD HH:MM' 字符串。"""
    if isinstance(val, datetime):
        return val
    if isinstance(val, str) and len(val) >= 16:
        try:
            return datetime.strptime(val[:16], '%Y-%m-%d %H:%M')
        except ValueError:
            pass
    return datetime.min


def create_empty_main(workspace, main_file, city_name):
    """首次运行：自动创建空主文件（12 列表头），零代码首次运行不中断。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '采购公告'
    # 第 1 行标题、第 2 行汇总行占位、第 3 行表头
    ws.cell(1, 1, f'{city_name}政府采购公告')
    ws.cell(1, 1).font = Font(bold=True, size=14)
    ws.cell(2, 1, f'采集时间：{datetime.now().strftime("%Y-%m-%d %H:%M")} | 数据来源：浙江政府采购网 | 共 0 条')
    for c, h in enumerate(ALL_COLUMNS, 1):
        cell = ws.cell(HEADER_ROW, c, h)
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal='center')
    for col, w in COL_WIDTHS.items():
        ws.column_dimensions[col].width = w
    wb.save(main_file)
    wb.close()
    print(f'主文件不存在，已自动创建空主文件：{main_file.name}')


def merge_new_files(workspace_dir, city_name='丽水市'):
    """Merge all {city}_采购公告_*.xlsx in workspace into main file, deduplicating by URL column"""
    workspace = Path(workspace_dir).resolve()
    city_base = city_name[:-1] if city_name.endswith('市') else city_name
    main_file = workspace / f'{city_base}政府采购公告.xlsx'

    if not workspace.exists():
        sys.exit(f'错误：工作目录不存在：{workspace}')

    # Excel lock file check
    lock_file = workspace / f'~${city_base}政府采购公告.xlsx'
    if lock_file.exists():
        sys.exit(f'错误：主文件被 Excel 打开，无法写入！\n请先关闭 Excel：\n  {main_file}\n锁文件：\n  {lock_file}')

    # 首次运行自动创建空主文件（12 列表头），保证零代码首次运行不中断
    if not main_file.exists():
        create_empty_main(workspace, main_file, city_name)

    # Find new batch files（按地市前缀匹配）
    pattern = str(workspace / f'{city_base}_采购公告_*.xlsx')
    batch_files = sorted(glob.glob(pattern))

    if not batch_files:
        print("No new batch files found.")
        return

    print(f"Found {len(batch_files)} batch file(s):")
    for f in batch_files:
        print(f"  {os.path.basename(f)}")

    # Load main file
    print(f"\nLoading main file: {main_file.name}")
    wb_main = openpyxl.load_workbook(main_file)
    ws_main = wb_main.active

    main_col = _col_index(ws_main, HEADER_ROW)
    url_col = main_col.get('地址链接URL')
    scr_col = main_col.get(STATUS_SCREENED_COL)
    jdg_col = main_col.get(STATUS_JUDGED_COL)
    if not url_col:
        wb_main.close()
        sys.exit(f'错误：主文件缺少"地址链接URL"列，请检查主文件格式：{main_file.name}')

    # Collect existing URLs for dedup
    existing_urls = set()
    for row in ws_main.iter_rows(min_row=FIRST_DATA_ROW, values_only=True):
        if row and len(row) >= url_col and row[url_col - 1]:
            existing_urls.add(str(row[url_col - 1]))

    print(f"  Existing records: {ws_main.max_row - (FIRST_DATA_ROW - 1)}")
    print(f"  Existing unique URLs: {len(existing_urls)}")

    new_count = 0
    skip_count = 0

    for batch_file in batch_files:
        wb_batch = openpyxl.load_workbook(batch_file)
        ws_batch = wb_batch.active
        batch_col = _col_index(ws_batch, 1)

        batch_total = 0
        batch_new = 0

        for row in ws_batch.iter_rows(min_row=2, values_only=True):
            if not row or not row[1]:  # col 0 is 序号, col 1 is 公告时间
                continue
            batch_total += 1

            # 批次文件列：0=序号，1=公告时间，…，10=地址链接URL（按列名定位更稳健）
            b_url_col = batch_col.get('地址链接URL')
            url = row[b_url_col - 1] if b_url_col and len(row) >= b_url_col else ''

            if url and str(url) in existing_urls:
                skip_count += 1
                continue

            # Append to main, skipping batch col 0 (序号)
            new_row = ws_main.max_row + 1
            for c in range(1, len(DATA_COLUMNS) + 1):
                col_name = DATA_COLUMNS[c - 1]
                b_col = batch_col.get(col_name)
                if b_col is None or len(row) < b_col:
                    val = ''
                else:
                    val = row[b_col - 1]
                if col_name == '地址链接URL' and val:
                    cell = ws_main.cell(new_row, c)
                    cell.value = val
                    cell.hyperlink = val
                    cell.style = 'Hyperlink'
                else:
                    ws_main.cell(new_row, c, val)

            # 状态列：新增记录置「新增待初筛 / 空」（方案 4.3）
            if scr_col is not None:
                ws_main.cell(new_row, scr_col, '新增待初筛')
            if jdg_col is not None:
                ws_main.cell(new_row, jdg_col, '')

            if url:
                existing_urls.add(str(url))
            new_count += 1
            batch_new += 1

        wb_batch.close()
        print(f"  {os.path.basename(batch_file)}: {batch_total} total, {batch_new} new")

    # Sort all data rows by 公告时间 DESC（搬运 1–12 列，状态列随行走）
    data_rows = []
    for r in range(FIRST_DATA_ROW, ws_main.max_row + 1):
        row_data = [ws_main.cell(r, c).value for c in range(1, len(ALL_COLUMNS) + 1)]
        if row_data[0]:
            data_rows.append(row_data)
    data_rows.sort(key=lambda r: _sort_key(r[0]), reverse=True)

    # Clear and write back sorted, normalizing time column to string without seconds
    for r in range(FIRST_DATA_ROW, ws_main.max_row + 1):
        for c in range(1, len(ALL_COLUMNS) + 1):
            ws_main.cell(r, c).value = None
    for i, row_data in enumerate(data_rows):
        r = FIRST_DATA_ROW + i
        for c in range(1, len(ALL_COLUMNS) + 1):
            val = row_data[c - 1]
            if c == 1:
                val = _normalize_time(val)
            if c == url_col and val:  # URL column -> hyperlink
                cell = ws_main.cell(r, c)
                cell.value = val
                cell.hyperlink = val
                cell.style = 'Hyperlink'
            else:
                ws_main.cell(r, c, val)

    # Time column stored as text to avoid Excel showing seconds
    for r in range(FIRST_DATA_ROW, FIRST_DATA_ROW + len(data_rows)):
        ws_main.cell(r, 1).number_format = '@'

    # Update the summary row (row 2)
    total_records = len(data_rows)
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M')
    ws_main.cell(2, 1).value = f'采集时间：{now_str} | 数据来源：浙江政府采购网 | 共 {total_records} 条'

    wb_main.save(main_file)
    wb_main.close()

    print(f"\n{'='*60}")
    print(f"Merge complete!")
    print(f"  New records: {new_count}")
    print(f"  Skipped (duplicates): {skip_count}")
    print(f"  Total records in main file: {total_records}")
    print(f"{'='*60}")

    # Clean up batch files
    for f in batch_files:
        try:
            os.remove(f)
            print(f"  Removed: {os.path.basename(f)}")
        except:
            pass


def _ensure_multi_sheet(wb, sheet_title, city_name):
    """确保多 Sheet 主文件目标 Sheet 存在且为 15 列标准结构（标题行/汇总行/表头行）。"""
    if sheet_title in wb.sheetnames:
        ws = wb[sheet_title]
        # 检查表头是否齐全（表头行自动探测：前 5 行内含「采购项目名称」的行）
        header_row = None
        for r in range(1, min(ws.max_row, 5) + 1):
            vals = [str(ws.cell(r, c).value or '').strip() for c in range(1, ws.max_column + 1)]
            if '采购项目名称' in vals:
                header_row = r
                break
        if header_row is None:
            # 旧格式 Sheet（表头在第 1 行）：重建为标准 15 列结构
            ws.cell(1, 1, f'{city_name}政府采购公告')
            ws.cell(1, 1).font = Font(bold=True, size=14)
            ws.cell(2, 1, f'采集时间：{datetime.now().strftime("%Y-%m-%d %H:%M")} | 数据来源：浙江政府采购网 | 共 0 条')
            for c, h in enumerate(MULTI_MAIN_COLUMNS, 1):
                cell = ws.cell(HEADER_ROW, c, h)
                cell.font = Font(bold=True, size=10)
                cell.alignment = Alignment(horizontal='center')
            for col, w in MULTI_COL_WIDTHS.items():
                ws.column_dimensions[col].width = w
        else:
            # 已有表头：补齐缺失列（旧 8/10 列格式 → 15 列）
            col_of = _col_index(ws, header_row)
            new_col = ws.max_column + 1
            for name in MULTI_MAIN_COLUMNS:
                if name not in col_of:
                    cell = ws.cell(header_row, new_col, name)
                    cell.font = Font(bold=True, size=10)
                    col_of[name] = new_col
                    new_col += 1
            for col, w in MULTI_COL_WIDTHS.items():
                if col not in ws.column_dimensions or not ws.column_dimensions[col].width:
                    ws.column_dimensions[col].width = w
        return ws
    ws = wb.create_sheet(title=sheet_title)
    ws.cell(1, 1, f'{city_name}政府采购公告')
    ws.cell(1, 1).font = Font(bold=True, size=14)
    ws.cell(2, 1, f'采集时间：{datetime.now().strftime("%Y-%m-%d %H:%M")} | 数据来源：浙江政府采购网 | 共 0 条')
    for c, h in enumerate(MULTI_MAIN_COLUMNS, 1):
        cell = ws.cell(HEADER_ROW, c, h)
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal='center')
    for col, w in MULTI_COL_WIDTHS.items():
        ws.column_dimensions[col].width = w
    return ws


def merge_to_multi_main(workspace_dir, main_file_path, city_name='丽水市'):
    """
    将 {city}_采购公告_*.xlsx 批次合并进多 Sheet 主文件（政府采购公告.xlsx）对应地市 Sheet。

    - Sheet 名 = 地市名（如"丽水"）；URL 去重、公告时间倒序；
    - 新增行状态：初筛状态=新增待初筛、研判状态=空；
    - 目标 Sheet 缺列自动补齐为 15 列（10 数据列 + 初筛状态/研判状态 + 初筛结果/初筛原因/相关性(命中词)）。
    """
    workspace = Path(workspace_dir).resolve()
    city_base = city_name[:-1] if city_name.endswith('市') else city_name
    main_file = Path(main_file_path).resolve()
    sheet_title = city_base

    if not workspace.exists():
        sys.exit(f'错误：工作目录不存在：{workspace}')
    if not main_file.exists():
        sys.exit(f'错误：主文件不存在：{main_file}')

    # Excel lock file check
    lock_file = main_file.parent / f'~${main_file.name}'
    if lock_file.exists():
        sys.exit(f'错误：主文件被 Excel 打开，无法写入！\n请先关闭 Excel：\n  {main_file}\n锁文件：\n  {lock_file}')

    # Find new batch files（按地市前缀匹配）
    pattern = str(workspace / f'{city_base}_采购公告_*.xlsx')
    batch_files = sorted(glob.glob(pattern))
    if not batch_files:
        print("No new batch files found.")
        return

    print(f"Found {len(batch_files)} batch file(s):")
    for f in batch_files:
        print(f"  {os.path.basename(f)}")

    wb_main = openpyxl.load_workbook(main_file)
    ws_main = _ensure_multi_sheet(wb_main, sheet_title, city_name)

    # 定位主 Sheet 表头（可能被 _ensure 重建或补齐）
    header_row = None
    for r in range(1, min(ws_main.max_row, 5) + 1):
        vals = [str(ws_main.cell(r, c).value or '').strip() for c in range(1, ws_main.max_column + 1)]
        if '采购项目名称' in vals:
            header_row = r
            break
    if header_row is None:
        wb_main.close()
        sys.exit(f'错误：主文件 Sheet「{sheet_title}」未找到表头行，请检查主文件格式。')
    main_col = _col_index(ws_main, header_row)
    url_col = main_col.get('地址链接URL')
    scr_col = main_col.get(STATUS_SCREENED_COL)
    jdg_col = main_col.get(STATUS_JUDGED_COL)
    if not url_col:
        wb_main.close()
        sys.exit(f'错误：主文件 Sheet「{sheet_title}」缺少"地址链接URL"列，请检查主文件格式。')

    # 已有 URL 集合（去重用）
    existing_urls = set()
    for row in ws_main.iter_rows(min_row=header_row + 1, values_only=True):
        if row and len(row) >= url_col and row[url_col - 1]:
            existing_urls.add(str(row[url_col - 1]))

    print(f"  Existing records: {ws_main.max_row - header_row}")
    print(f"  Existing unique URLs: {len(existing_urls)}")

    new_count = 0
    skip_count = 0

    for batch_file in batch_files:
        wb_batch = openpyxl.load_workbook(batch_file)
        ws_batch = wb_batch.active
        batch_col = _col_index(ws_batch, 1)
        batch_total = 0
        batch_new = 0
        for row in ws_batch.iter_rows(min_row=2, values_only=True):
            if not row or not row[1]:
                continue
            batch_total += 1
            b_url_col = batch_col.get('地址链接URL')
            url = row[b_url_col - 1] if b_url_col and len(row) >= b_url_col else ''
            if url and str(url) in existing_urls:
                skip_count += 1
                continue
            new_row = ws_main.max_row + 1
            for name in MULTI_MAIN_COLUMNS:
                c = main_col.get(name)
                if c is None:
                    continue
                b_col = batch_col.get(name)
                val = row[b_col - 1] if b_col is not None and len(row) >= b_col else ''
                if name == '地址链接URL' and val:
                    cell = ws_main.cell(new_row, c)
                    cell.value = val
                    cell.hyperlink = val
                    cell.style = 'Hyperlink'
                else:
                    ws_main.cell(new_row, c, val)
            # 状态列：新增记录置「新增待初筛 / 空」
            if scr_col is not None:
                ws_main.cell(new_row, scr_col, '新增待初筛')
            if jdg_col is not None:
                ws_main.cell(new_row, jdg_col, '')
            if url:
                existing_urls.add(str(url))
            new_count += 1
            batch_new += 1
        wb_batch.close()
        print(f"  {os.path.basename(batch_file)}: {batch_total} total, {batch_new} new")

    # 按公告时间倒序重排（搬运全部 15 列，状态随行走）
    data_rows = []
    for r in range(header_row + 1, ws_main.max_row + 1):
        row_data = [ws_main.cell(r, c).value for c in range(1, len(MULTI_MAIN_COLUMNS) + 1)]
        if row_data[0]:
            data_rows.append(row_data)
    data_rows.sort(key=lambda r: _sort_key(r[0]), reverse=True)

    for r in range(header_row + 1, ws_main.max_row + 1):
        for c in range(1, len(MULTI_MAIN_COLUMNS) + 1):
            ws_main.cell(r, c).value = None
    for i, row_data in enumerate(data_rows):
        r = header_row + 1 + i
        for c in range(1, len(MULTI_MAIN_COLUMNS) + 1):
            val = row_data[c - 1]
            if c == 1:
                val = _normalize_time(val)
            if c == url_col and val:
                cell = ws_main.cell(r, c)
                cell.value = val
                cell.hyperlink = val
                cell.style = 'Hyperlink'
            else:
                ws_main.cell(r, c, val)
    for r in range(header_row + 1, header_row + 1 + len(data_rows)):
        ws_main.cell(r, 1).number_format = '@'

    total_records = len(data_rows)
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M')
    ws_main.cell(2, 1).value = f'采集时间：{now_str} | 数据来源：浙江政府采购网 | 共 {total_records} 条'

    wb_main.save(main_file)
    wb_main.close()

    print(f"\n{'='*60}")
    print(f"Merge complete!（多 Sheet 主文件：{main_file.name} / Sheet：{sheet_title}）")
    print(f"  New records: {new_count}")
    print(f"  Skipped (duplicates): {skip_count}")
    print(f"  Total records in Sheet「{sheet_title}」: {total_records}")
    print(f"{'='*60}")

    for f in batch_files:
        try:
            os.remove(f)
            print(f"  Removed: {os.path.basename(f)}")
        except:
            pass


if __name__ == '__main__':
    args = get_args()
    if args.main_file:
        merge_to_multi_main(args.workspace, args.main_file, args.city)
    else:
        merge_new_files(args.workspace, args.city)
