# -*- coding: utf-8 -*-
"""
collect_round.py —— 增量采集（多 Sheet 主文件，按各城市 Sheet 最后公告时间起算）

用法：
    python scripts/collect_round.py --workspace "清单路径" [--cities 丽水市,杭州市] [--chunk-days 2]

流程（每城市）：
1. 读多 Sheet 主文件（政府采购公告.xlsx）对应城市 Sheet，取最大「公告时间」作为起算日期
   （起算日包含在内：当天下午新发布的公告也能补到；无数据时默认最近 3 天）；
2. 按 chunk-days（默认 2 天）分批调用 fetch_zfcg.py（--range/--date）；
3. merge_to_main.py --main-file 合并入多 Sheet 主文件（URL 去重、自动标新增待初筛）。

幂等：重复采集由 URL 去重兜底，可随时重跑。
"""
from __future__ import annotations

import argparse
import json
import subprocess
import sys
import os
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
ALL_CITIES = ["杭州市", "宁波市", "温州市", "嘉兴市", "湖州市", "绍兴市",
              "金华市", "衢州市", "舟山市", "台州市", "丽水市"]


def _sheet_info(main_path: Path, city: str):
    """读取多 Sheet 主文件中城市 Sheet 的最大公告时间与已入库 URL 集合。
    返回 (max_dt 或 None, urls 集合)。"""
    sheet_title = city[:-1] if city.endswith("市") else city
    try:
        wb = openpyxl.load_workbook(main_path)
    except Exception:
        return None, set()
    if sheet_title not in wb.sheetnames:
        wb.close()
        return None, set()
    ws = wb[sheet_title]
    hr = None
    for r in range(1, min(ws.max_row, 5) + 1):
        vals = [str(ws.cell(r, c).value or "").strip() for c in range(1, ws.max_column + 1)]
        if "公告时间" in vals:
            hr = r
            break
    if hr is None:
        wb.close()
        return None, set()
    col_of = {}
    for c in range(1, ws.max_column + 1):
        name = str(ws.cell(hr, c).value or "").strip()
        if name:
            col_of.setdefault(name, c)
    time_col = col_of.get("公告时间")
    url_col = col_of.get("地址链接URL")
    if time_col is None:
        wb.close()
        return None, set()
    max_dt = None
    urls = set()
    for r in range(hr + 1, ws.max_row + 1):
        v = ws.cell(r, time_col).value
        if isinstance(v, datetime):
            max_dt = max(max_dt, v) if max_dt else v
        elif isinstance(v, str):
            for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d"):
                try:
                    d = datetime.strptime(v[:16], fmt)
                    max_dt = max(max_dt, d) if max_dt else d
                    break
                except ValueError:
                    continue
        if url_col is not None:
            u = str(ws.cell(r, url_col).value or "").strip()
            if u:
                urls.add(u.lower())
    wb.close()
    return max_dt, urls


def _date_range(start: datetime, end: datetime, chunk_days: int):
    """按 chunk_days 把 [start, end] 拆成日期区间。"""
    chunks = []
    cur = start
    while cur <= end:
        nxt = min(cur + timedelta(days=chunk_days - 1), end)
        chunks.append((cur.date(), nxt.date()))
        cur = nxt + timedelta(days=1)
    return chunks


def run(cmd, **kw):
    print("  $ " + " ".join(cmd))
    r = subprocess.run(cmd, cwd=str(ROOT), **kw)
    return r.returncode


def main() -> int:
    parser = argparse.ArgumentParser(description="增量采集（多 Sheet 主文件）")
    parser.add_argument("--workspace", required=True, help="清单路径（主文件与批次文件所在）")
    parser.add_argument("--cities", default=None, help="逗号分隔地市（默认全部 11 地市）")
    parser.add_argument("--main-file", default=None, help="多 Sheet 主文件路径（默认 workspace/政府采购公告.xlsx）")
    parser.add_argument("--chunk-days", type=int, default=2, help="每批天数（默认 2）")
    parser.add_argument("--max-days", type=int, default=10, help="起算日最多回溯天数（默认 10，防止 Sheet 损坏时拉太多）")
    args = parser.parse_args()

    ws = Path(args.workspace).resolve()
    main_file = Path(args.main_file).resolve() if args.main_file else ws / "政府采购公告.xlsx"
    if not main_file.exists():
        print("主文件不存在：%s" % main_file)
        return 1
    cities = [c.strip() for c in (args.cities or "").split(",") if c.strip()] or ALL_CITIES
    today = datetime.now()

    print("主文件：%s" % main_file)
    print("地市：%s" % "、".join(cities))
    for i, city in enumerate(cities, 1):
        print("\n########## [%d/%d] %s ##########" % (i, len(cities), city))
        max_dt, existing_urls = _sheet_info(main_file, city)
        if max_dt is None:
            print("  ↳ Sheet 无数据或不存在，按最近 3 天起算")
            start = today - timedelta(days=2)
        else:
            # 起算日含当日：当天下午新发布公告也能补到
            start = max_dt.replace(hour=0, minute=0, second=0, microsecond=0)
            if (today.date() - start.date()).days > args.max_days:
                start = today - timedelta(days=args.max_days - 1)
                print("  ↳ 最后公告日期过早，截断为最近 %d 天" % args.max_days)
        # 已入库 URL 清单落临时文件，供采集器详情抓取前过滤（跳过重复公告，大幅提速）
        existing_file = ws / ("_existing_urls_%s.json" % city[:-1] if city.endswith("市") else city)
        with open(existing_file, "w", encoding="utf-8") as f:
            json.dump({"urls": sorted(existing_urls)}, f, ensure_ascii=False)
        print("  ↳ 已入库 URL %d 条，采集时将跳过其详情抓取" % len(existing_urls))
        end = today
        chunks = _date_range(start, end, args.chunk_days)
        print("  ↳ 起算：%s → 今天（共 %d 天，分 %d 批）" % (
            start.date(), (end.date() - start.date()).days + 1, len(chunks)))
        failed = False
        for ci, (c_start, c_end) in enumerate(chunks, 1):
            if c_start == c_end:
                fetch_cmd = [sys.executable, "scripts/fetch_zfcg.py", "--city", city,
                             "--date", c_start.strftime("%Y-%m-%d"), "--workspace", str(ws),
                             "--existing-urls", str(existing_file)]
            else:
                fetch_cmd = [sys.executable, "scripts/fetch_zfcg.py", "--city", city,
                             "--range", c_start.strftime("%Y-%m-%d"), c_end.strftime("%Y-%m-%d"),
                             "--workspace", str(ws), "--existing-urls", str(existing_file)]
            print("  ↳ 第 %d/%d 批：%s ~ %s" % (ci, len(chunks), c_start, c_end))
            rc = run(fetch_cmd)
            if rc != 0:
                print("  ✗ 第 %d 批采集失败（rc=%d），继续下一批" % (ci, rc))
                failed = True
        try:
            existing_file.unlink(missing_ok=True)
        except Exception:
            pass
        if failed:
            print("  ↳ 存在失败批次，跳过合并（批次文件保留，重跑本命令可续）")
            continue
        merge_cmd = [sys.executable, "scripts/merge_to_main.py", "--city", city,
                     "--main-file", str(main_file), "--workspace", str(ws)]
        rc = run(merge_cmd)
        if rc != 0:
            print("  ✗ 合并失败（rc=%d），批次文件保留在工作区，稍后可重跑" % rc)
    print("\n########## 增量采集完成 ##########")
    return 0


if __name__ == "__main__":
    sys.exit(main())
