# -*- coding: utf-8 -*-
"""
政府采购公告数据采集脚本 (ZFCG Data Collector)
==============================================
Version: 2.0
从浙江政府采购网 (zfcg.czt.zj.gov.cn) 采集指定地区的采购公告数据，
通过 API + HTML 解析 + 正则提取 + 交叉引用的方式获取结构化信息，
最终输出 Excel 文件。

用法：
  python fetch_zfcg.py --workspace C:\\path\\to\\workspace             # 自动模式（分析主文件缺失日期）
  python fetch_zfcg.py --date 2026-06-22 --workspace C:\\path         # 手动指定单日
  python fetch_zfcg.py --range 2026-06-20 2026-06-22 --workspace C:\\path  # 手动指定范围
  python fetch_zfcg.py --city 杭州市 --workspace C:\\path             # 指定地市（默认丽水市）
  python fetch_zfcg.py --date 2026-06-22 --workspace . --verify      # 采集后运行验证

依赖：pip install openpyxl
"""
import json
import openpyxl
import urllib.request
import urllib.parse
import re
import time
import html
import sys
import os
from openpyxl.styles import Font, Alignment
from datetime import datetime, timedelta
from collections import defaultdict
import argparse
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8')

# ========================================================================
# Configuration — set via --workspace CLI argument (see main())
# ========================================================================
MAIN_FILE = None  # Will be set in main() based on --workspace
WORKSPACE_DIR = None  # Will be set in main() based on --workspace（批次文件落盘位置）

# 技能根目录（本文件位于 scripts/ 下）与地市配置文件
ROOT_DIR = Path(__file__).resolve().parent.parent
DISTRICTS_FILE = ROOT_DIR / 'references' / 'districts.json'

# 目标地市（--city 指定，默认"丽水市"保持向后兼容）；CITY_BASE = 去"市"后缀，用于文件名
CITY_NAME = '丽水市'
CITY_BASE = '丽水'

# 公告类型 → categoryCode 映射（浙江政府采购网）
CATEGORY_CODES = {
    '采购意向': '110-175885',
    '意见征询': '110-246839',
    '采购项目公告': '110-978863',
    '更正公告': '110-943756',
    '采购结果公告': '110-900461',
    '采购合同公告': '110-567245',
    '履约验收公告': '110-198517',
    '电子卖场公告': '110-341071',
    '框架协议公告': '110-978920',
}

# pathName → 标准类型 映射
PATH_TO_CATEGORY = {
    '采购意向公开': '采购意向',
    '采购文件需求公示': '意见征询',
    '单一来源采购公示': '意见征询',
    '资格预审公告': '采购项目公告',
    '招标公告': '采购项目公告',
    '非招标公告': '采购项目公告',
    '更正公告': '更正公告',
    '中标（成交）结果公告': '采购结果公告',
    '废标公告': '采购结果公告',
    '采购合同公告': '采购合同公告',
    '合同变更公告': '采购合同公告',
    '履约验收公告': '履约验收公告',
    '其他电子卖场公告': '电子卖场公告',
    '框架协议公告': '框架协议公告',
}

# 目标地区区县编码 → 名称 映射（运行时按 --city 从 references/districts.json 加载，见 load_districts）
# 格式: {districtCode: '区县名'}
DISTRICT_MAP = {}

# 数据完整性阈值
MIN_RECORDS_PER_DAY = 30   # 每天最低记录数
MIN_LATEST_HOUR = 17        # 最晚记录小时数

# 构建快速查找集合（在 main() 中按 --city 动态填充）
TARGET_DISTRICTS = set()
TARGET_DISTRICT_CODES = set()


def load_districts(city_name):
    """按城市名从 references/districts.json 加载区县映射（静态配置，技能内置）。

    返回 {districtCode: '区县名'}；城市不存在或配置文件缺失/损坏时中文提示并退出。
    """
    if not DISTRICTS_FILE.exists():
        sys.exit(f'错误：未找到地市配置文件：{DISTRICTS_FILE}\n请确认技能目录 references/districts.json 存在。')
    try:
        with open(DISTRICTS_FILE, encoding='utf-8') as f:
            data = json.load(f)
    except Exception as e:
        sys.exit(f'错误：地市配置文件读取失败：{DISTRICTS_FILE}\n详细信息：{e}')
    cities = (data or {}).get('cities', {})
    city = cities.get(city_name)
    if not city:
        sys.exit(f'错误：不支持的地市：{city_name}\n支持的地市：{"、".join(cities.keys())}')
    return dict(city.get('districts', {}))

# 供应商名称合法后缀（末尾必须包含其中之一）
SUPPLIER_SUFFIXES = (
    '公司|中心|事务所|集团|医院|院校|大学|研究院|书院|学院|'
    '合作社|工程队|协会|勘查院|农资店|服务部|家庭农场|商行|经营部|养殖场|'
    '客栈|工作室|实验室|配送有限公司|分公司'
)

# 采购人名称合法后缀
PURCHASER_SUFFIXES = '局|委|办|中心|院|处|部|所|站|队|园|校|厅|室|会|学'

# API 端点
LIST_API = 'https://zfcg.czt.zj.gov.cn/portal/category'
DETAIL_API = 'https://zfcg.czt.zj.gov.cn/portal/detail'

API_HEADERS = {
    'Content-Type': 'application/json',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
    'Accept': 'application/json, text/plain, */*',
    'Origin': 'https://zfcg.czt.zj.gov.cn',
    'Referer': 'https://zfcg.czt.zj.gov.cn/site/category?parentId=600007&childrenCode=ZcyAnnouncement',
}

# ========================================================================
# Text & Extraction Utilities
# ========================================================================

def strip_html(s):
    """Remove HTML tags and normalize whitespace."""
    s = re.sub(r'<style[^>]*>.*?</style>', '', s, flags=re.DOTALL)
    s = re.sub(r'<[^>]+>', ' ', s)
    s = html.unescape(s)
    s = re.sub(r'&nbsp;', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def extract_budget(text):
    """Extract budget amount (预算金额) from text. Returns float or None."""
    norm = re.sub(r'\s+', ' ', text)
    patterns = [
        (r'预算金额[（(]元[)）][：:\s]*([\d,.]+)', False),
        (r'预算金额[：:\s]*([\d,.]+)\s*元', False),
        (r'预算总价[（(]元[)）][：:\s]*([\d,.]+)', False),
        (r'预算总价[：:\s]*([\d,.]+)\s*元', False),
        (r'采购预算[（(]元[)）][：:\s]*([\d,.]+)', False),
        (r'采购预算[：:\s]*([\d,.]+)\s*元', False),
        (r'项目预算[：:\s]*([\d,.]+)\s*元', False),
        (r'预算金额[：:\s]*人民币\s*([\d,.]+)', False),
        (r'最高限价[（(]元[)）][：:\s]*([\d,.]+)', False),
        (r'最高限价[：:\s]*([\d,.]+)\s*元', False),
        (r'控制价[：:\s]*([\d,.]+)\s*元', False),
        (r'单一来源采购限额[：:\s]*([\d,.]+)\s*万元', True),
        (r'单一来源采购限额[：:\s]*([\d,.]+)\s*元', False),
        (r'预算[（(]万元[)）][：:\s]*([\d,.]+)', True),
        (r'预算[：:\s]*([\d,.]+)\s*元', False),
        (r'预算[：:\s]*([\d,.]+)\s*万元', True),
        (r'￥\s*([\d,]+(?:\.\d{1,2})?)', False),
    ]
    for p, is_wan in patterns:
        m = re.search(p, norm)
        if m:
            val = m.group(1).replace(',', '')
            try:
                v = float(val)
                if is_wan or '万元' in norm[max(0, m.start() - 10):m.end() + 10]:
                    v *= 10000
                if v > 0:
                    return v
            except ValueError:
                pass
    return None


def extract_winning_amount(text):
    """Extract winning amount (中标金额) from text. Returns float or None."""
    norm = re.sub(r'\s+', ' ', text)

    # Multi-lot: sum all 总价 patterns
    multi = re.findall(r'总价[：:,，\s]*([\d,.]+)\s*[（(]元[)）]', norm)
    if len(multi) >= 2:
        total = 0
        for val in multi:
            try:
                total += float(val.replace(',', ''))
            except ValueError:
                pass
        if total > 0:
            return total

    patterns = [
        r'中标[（(]成交[)）]金额[（(]元[)）][：:\s]*([\d,.]+)',
        r'中标金额[（(]元[)）][：:\s]*([\d,.]+)',
        r'成交金额[（(]元[)）][：:\s]*([\d,.]+)',
        r'中标[（(]成交[)）]金额[：:\s]*([\d,.]+)\s*元',
        r'合同金额[（(]元[)）][：:\s]*([\d,.]+)',
        r'合同金额[：:\s]*([\d,.]+)',
        r'总中标金额[：:\s]*([\d,.]+)',
        r'成交总金额[：:\s]*([\d,.]+)',
        r'中标总价[：:\s]*([\d,.]+)',
        r'总价报价[（(]元[)）][：:\s]*([\d,.]+)',
        r'总价报价[：:\s]*([\d,.]+)\s*[（(]元[)）]',
        r'总价报价[：:\s]*([\d,.]+)\s*元',
        r'总报价[（(]元[)）][：:\s]*([\d,.]+)',
        r'总报价[（(][^)）]*[)）][：:\s]*([\d,.]+)',
        r'总报价[：:\s]*([\d,.]+)\s*[（(]元[)）]',
        r'总报价[：:\s]*([\d,.]+)\s*元',
        r'总价[（(][^)）]*[)）][：:\s]*([\d,.]+)',
        r'总价[：:,，\s]*([\d,.]+)\s*[（(]元[)）]',
        r'总价[：:,，\s]*([\d,.]+)\s*元',
        r'投标报价[（(]元[)）][：:\s]*([\d,.]+)',
        r'投标报价[：:\s]*([\d,.]+)\s*[（(]元[)）]',
        r'投标报价[：:\s]*([\d,.]+)\s*元',
        r'投标价[：:\s]*([\d,.]+)\s*[（(]元[)）]',
        r'报价[：:]\s*([\d,.]+)\s*[（(]元[)）]',
        r'投标单价[（(][^)）]*[)）][：:\s]*([\d,.]+)\s*[（(]元[)）]',
        r'单价[：:\s]*([\d,.]+)\s*[（(]元[)）]',
        r'设备报价[：:\s]*([\d,.]+)\s*[（(]元[)）]',
        r'合价[（(][^)）]*[)）][：:\s]*([\d,.]+)',
        r'合价[：:\s]*([\d,.]+)\s*[（(]元[)）]',
    ]
    for p in patterns:
        m = re.search(p, norm)
        if m:
            val = m.group(1).replace(',', '')
            try:
                return float(val)
            except ValueError:
                pass

    # Percentage pricing detection: "投标结算比例：10（%）" or "试剂报价：7（%）"
    pct_match = re.search(r'(?:结算比例|结算折扣|费率|下浮率)[：:\s]*([\d,.]+)\s*[（(]%[)）]', norm)
    if pct_match:
        return None  # Percentage pricing — deliberately return None, not a number

    # Generic fallback: find any "（元）" preceded by a number
    # Look in the winning section first
    win_sec = norm[:8000]
    m_ws = re.search(r'(?:中标|成交|采购结果)', win_sec)
    if not m_ws:
        m_ws = re.search(r'(?:三、)', win_sec)
    if m_ws:
        win_sec = win_sec[m_ws.start():m_ws.start()+3000]
    
    # Find the LAST amount+（元）that is followed by a company name
    amounts = list(re.finditer(r'([\d,.]+)\s*[（(]元[)）]', win_sec))
    for m_amt in reversed(amounts):
        after = win_sec[m_amt.end():m_amt.end()+100]
        if re.search(rf'(?:{SUPPLIER_SUFFIXES})', after):
            val = m_amt.group(1).replace(',', '')
            try:
                return float(val)
            except ValueError:
                pass
    
    return None


def extract_supplier(text):
    """Extract supplier name (供应商) from text. Returns str or None."""
    suffix_re = SUPPLIER_SUFFIXES
    norm = re.sub(r'\s+', ' ', text)

    # Multi-lot: extract all suppliers from 总价 lines
    multi = re.findall(
        rf'总价[：:,，\s]*[\d,.]+\s*[（(]元[)）]\s*([\u4e00-\u9fff（）()A-Za-z]+(?:{suffix_re}))',
        norm
    )
    if len(multi) >= 2:
        seen = set()
        unique = []
        for s in multi:
            if s not in seen:
                seen.add(s)
                unique.append(s)
        return ','.join(unique)

    # Joint bid detection
    m = re.search(
        rf'牵头供应商[：:]\s*([\u4e00-\u9fff（）()A-Za-z]+(?:{suffix_re})(?:[\u4e00-\u9fff（）()]*?'
        rf'(?:分公司|{suffix_re}))?)',
        norm
    )
    if m:
        leader = m.group(1).strip()
        idx = norm.find('投标联合体')
        if idx > -1:
            rest = norm[idx + len('投标联合体'):]
            for boundary in [
                '中标供应商地址', '供应商地址', '地址：', '四、',
                ' 1.', '省杭州市', '省丽水市', '省温州市', '省宁波市'
            ]:
                bi = rest.find(boundary)
                if bi > 5:
                    rest = rest[:bi]
                    break
            rest = re.sub(r'[：:\s]+$', '', rest).strip()
            companies = re.findall(rf'([\u4e00-\u9fff（）()A-Za-z]+(?:{suffix_re}))', rest)
            if companies:
                return f"牵头：{leader}；联合体：{'、'.join(companies)}"
        return leader

    # Standard patterns
    patterns = [
        rf'中标供应商名称[：:\s]*([\u4e00-\u9fff（）()A-Za-z]+(?:{suffix_re})(?:[\u4e00-\u9fff（）()]*?(?:分公司|{suffix_re}))?)',
        rf'中标[（(]成交[)）]供应商[：:\s]*([\u4e00-\u9fff（）()A-Za-z]+(?:{suffix_re})(?:[\u4e00-\u9fff（）()]*?(?:分公司|公司|中心))?)',
        rf'成交供应商[：:\s]*([\u4e00-\u9fff（）()A-Za-z]+(?:{suffix_re})(?:[\u4e00-\u9fff（）()]*?(?:分公司|公司|中心))?)',
        rf'供应商[（(]乙方[)）][：:\s]*([\u4e00-\u9fff（）()A-Za-z]+(?:{suffix_re})(?:[\u4e00-\u9fff（）()]*?(?:分公司|公司|中心))?)',
        rf'入围供应商[：:\s]*([\u4e00-\u9fff（）()A-Za-z]+(?:{suffix_re})(?:[\u4e00-\u9fff（）()]*?(?:分公司|公司|中心))?)',
        rf'供应商名称[：:\s]*([\u4e00-\u9fff（）()A-Za-z]+(?:{suffix_re})(?:[\u4e00-\u9fff（）()]*?(?:分公司|公司|中心))?)',
    ]
    for p in patterns:
        m = re.search(p, text)
        if m:
            val = m.group(1).strip()
            if 2 < len(val) < 100:
                return val

    # Fallback: company name after amount patterns
    norm2 = re.sub(r'\s+', ' ', text)
    fallback_patterns = [
        rf'报价[：:]\s*[\d,.]+\s*[（(]元[)）]\s*([\u4e00-\u9fff（）()A-Za-z]+(?:{suffix_re})[)）]?)',
        rf'投标报价[：:\s]*[\d,.]+\s*[（(]元[)）]\s*([\u4e00-\u9fff（）()A-Za-z]+(?:{suffix_re})[)）]?)',
        rf'投标价[：:\s]*[\d,.]+\s*[（(]元[)）]\s*([\u4e00-\u9fff（）()A-Za-z]+(?:{suffix_re})[)）]?)',
        rf'总价[（(][^)）]*[)）][：:\s]*[\d,.]+\s*[（(]?元?[)）]?\s+([\u4e00-\u9fff（）()A-Za-z]+(?:{suffix_re})[)）]?)',
        rf'总价[：:,，\s]*[\d,.]+\s*[（(]元[)）]\s*([\u4e00-\u9fff（）()A-Za-z]+(?:{suffix_re})[)）]?)',
        rf'总价[：:\s]*[\d,.]+\s*[（(]%[)）]\s+([\u4e00-\u9fff（）()A-Za-z]+(?:{suffix_re})[)）]?)',
        rf'总价报价[：:\s]*[\d,.]+\s*[（(]元[)）]\s*([\u4e00-\u9fff（）()A-Za-z]+(?:{suffix_re})[)）]?)',
        rf'总报价[：:\s]*[\d,.]+\s*[（(]元[)）]\s*([\u4e00-\u9fff（）()A-Za-z]+(?:{suffix_re})[)）]?)',
        rf'总报价[（(][^)）]*[)）][：:\s]*[\d,.]+\s*[（(]?元?[)）]?\s*([\u4e00-\u9fff（）()A-Za-z]+(?:{suffix_re})[)）]?)',
        rf'总报价[（(]元[)）][：:\s]*[\d,.]+\s*[（(]?元?[)）]?\s*([\u4e00-\u9fff（）()A-Za-z]+(?:{suffix_re})[)）]?)',
        rf'中标[（(]成交[)）]金额[（(]元[)）][：:\s]*[\d,.]+\s*[（(]?元?[)）]?\s+([\u4e00-\u9fff（）()A-Za-z]+(?:{suffix_re})[)）]?)',
        rf'投标单价[（(][^)）]*[)）][：:\s]*[\d,.]+\s*[（(]元[)）]\s*([\u4e00-\u9fff（）()A-Za-z]+(?:{suffix_re})[)）]?)',
        rf'单价[：:\s]*[\d,.]+\s*[（(]元[)）]\s*([\u4e00-\u9fff（）()A-Za-z]+(?:{suffix_re})[)）]?)',
        rf'设备报价[：:\s]*[\d,.]+\s*[（(]元[)）][^。]{{0,50}}?([\u4e00-\u9fff（）()A-Za-z]+(?:{suffix_re})[)）]?)',
    ]
    for p in fallback_patterns:
        m = re.search(p, norm2)
        if m:
            val = m.group(1).strip()
            if 2 < len(val) < 100:
                return val

    # Generic catch-all: any amount + (元) followed by company name
    m = re.search(
        rf'[\d,.]+\s*[（(]元[)）]\s*([\u4e00-\u9fff（）()A-Za-z]+(?:{suffix_re})[)）]?)',
        norm2
    )
    if m:
        val = m.group(1).strip()
        if 2 < len(val) < 100:
            return val

    # Last-resort: find LAST (元) in winning section, get company after it
    # Handles multi-price rows like "总报价：123（元），...，456（元） 公司名"
    winning_sec = norm2[:5000]
    m_win = re.search(r'(?:中标|成交|采购结果)', winning_sec)
    if m_win:
        winning_sec = winning_sec[m_win.start():]
    last_yuan = [m for m in re.finditer(r'[（(]元[)）]', winning_sec)]
    if last_yuan:
        pos = last_yuan[-1].end()
        after = winning_sec[pos:pos+200]
        m_last = re.search(
            rf'([\u4e00-\u9fff（）()A-Za-z]+(?:{suffix_re})[)）]{0,2})(?:\s|，|。)',
            after.lstrip()
        )
        if m_last:
            val = m_last.group(1).strip()
            if 2 < len(val) < 100:
                return val

    return None


def extract_agency(text, cat_type=''):
    """Extract agency name (代理机构) from text. Returns str or None."""
    if cat_type == '采购意向':
        return ''

    suffix_re = '公司|中心|事务所|代理'
    norm = re.sub(r'\s+', ' ', text[:5000])

    methods = [
        rf'([\u4e00-\u9fff（）()A-Za-z]+(?:{suffix_re}))\s+受\s+[\u4e00-\u9fff（）()]+(?:的)?\s*委托',
        rf'采购代理机构\s*信息\s*名\s*称[：:]\s*([\u4e00-\u9fff（）()A-Za-z]+(?:{suffix_re}))',
        rf'(?:接收)?代理\s*机构[：:]\s*([\u4e00-\u9fff（）()A-Za-z]+(?:{suffix_re}))',
        rf'采购代理机构.*?名\s*称[：:]\s*([\u4e00-\u9fff（）()A-Za-z]+(?:{suffix_re}))',
    ]
    for p in methods:
        m = re.search(p, norm)
        if m:
            val = m.group(1).strip()
            if 3 < len(val) < 60:
                return val
    return None


def extract_purchaser(text):
    """Extract purchaser name (采购人) from text. Returns str or None."""
    suffix_re = PURCHASER_SUFFIXES
    norm = re.sub(r'\s+', ' ', text[:3000])
    patterns = [
        rf'采购人[（(]甲方[)）][：:]\s*([\u4e00-\u9fff（）()]+(?:{suffix_re}))',
        rf'采购人[：:]\s*([\u4e00-\u9fff（）()]+(?:{suffix_re}))',
        rf'采购人.*?名\s*称[：:]\s*([\u4e00-\u9fff（）()]+(?:{suffix_re}))',
        rf'甲方[：:]\s*([\u4e00-\u9fff（）()]+(?:{suffix_re}))',
        rf'受\s*([\u4e00-\u9fff（）()]+(?:{suffix_re}|分局))\s*的?\s*委托',
        r'采购人信息.*?名\s*称[：:]\s*([\u4e00-\u9fff（）()A-Za-z]+)',
    ]
    for p in patterns:
        m = re.search(p, norm)
        if m:
            val = m.group(1).strip()
            if 2 < len(val) < 80:
                return val
    return None


# ========================================================================
# Stage 0: Date Analysis
# ========================================================================

def analyze_main_file():
    """Analyze the main Excel file and identify incomplete dates."""
    print("=== Stage 0: Analyze Data Completeness ===\n")

    if not os.path.exists(MAIN_FILE):
        print(f"  Main file '{MAIN_FILE}' not found. Defaulting to last 3 days.")
        today = datetime.now()
        return [(today - timedelta(days=i)).strftime('%Y-%m-%d') for i in range(2, -1, -1)]

    wb = openpyxl.load_workbook(MAIN_FILE)  # read_only=True removed due to Python 3.13 crash
    ws = wb.active

    date_stats = defaultdict(lambda: {'count': 0, 'latest_hour': 0, 'latest_time': ''})
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or not row[0]:
            continue
        dt_val = row[0]
        if isinstance(dt_val, datetime):
            date_str = dt_val.strftime('%Y-%m-%d')
            hour = dt_val.hour
            time_str = dt_val.strftime('%H:%M')
        else:
            dt_str = str(dt_val)
            if len(dt_str) >= 10:
                date_str = dt_str[:10]
                try:
                    hour = int(dt_str[11:13]) if len(dt_str) > 12 else 0
                    time_str = dt_str[11:16] if len(dt_str) > 12 else '00:00'
                except (ValueError, IndexError):
                    hour, time_str = 0, '00:00'
            else:
                continue
        date_stats[date_str]['count'] += 1
        if hour > date_stats[date_str]['latest_hour']:
            date_stats[date_str]['latest_hour'] = hour
            date_stats[date_str]['latest_time'] = time_str
    wb.close()

    today = datetime.now()
    today_str = today.strftime('%Y-%m-%d')
    current_hour = today.hour
    print(f"  Current time: {today.strftime('%Y-%m-%d %H:%M')}\n")

    print(f"  {'Date':<12} {'Records':>8} {'Latest':<8} {'Status':<10} {'Reason'}")
    print(f"  {'-'*60}")

    incomplete = []
    # 与 scripts/analyze_main.py、SKILL.md「近 10 天」口径一致（原为 7 天，口径不一致会漏补采）
    for i in range(10):
        check_date = (today - timedelta(days=i)).strftime('%Y-%m-%d')
        stats = date_stats.get(check_date)

        if stats is None:
            status, reason = "MISSING", "No data"
            incomplete.append(check_date)
        elif stats['count'] < MIN_RECORDS_PER_DAY:
            status, reason = "INCOMPLETE", f"Too few ({stats['count']}<{MIN_RECORDS_PER_DAY})"
            incomplete.append(check_date)
        elif stats['latest_hour'] < MIN_LATEST_HOUR:
            status, reason = "INCOMPLETE", f"No afternoon data (latest {stats['latest_hour']:02d}:xx)"
            incomplete.append(check_date)
        elif check_date == today_str and current_hour < 18:
            status, reason = "PENDING", "Today, may be incomplete"
            if stats['count'] < MIN_RECORDS_PER_DAY:
                incomplete.append(check_date)
        else:
            status, reason = "COMPLETE", ""

        latest = stats['latest_time'] if stats and stats['latest_time'] else '--:--'
        count = stats['count'] if stats else 0
        print(f"  {check_date:<12} {count:>8} {latest:<8} {status:<10} {reason}")

    if incomplete:
        fetch_dates = sorted(set(incomplete))
        print(f"\n  Dates to fetch: {', '.join(fetch_dates)}")
    else:
        yesterday = (today - timedelta(days=1)).strftime('%Y-%m-%d')
        fetch_dates = [yesterday]
        if current_hour >= 18:
            fetch_dates.append(today_str)
        print(f"\n  All data complete. Routine fetch: {', '.join(fetch_dates)}")

    if today_str in fetch_dates and current_hour < 18:
        print(f"\n  WARNING: Current time is {current_hour}:xx. Afternoon data may not be posted yet.")
        print(f"  Recommend re-collecting today's data after 18:00.")
    print()
    return fetch_dates


# ========================================================================
# Stage 1: List Data Fetching
# ========================================================================

def fetch_list_data(fetch_dates):
    """Fetch announcement list data from the API for the given dates."""
    print("=== Stage 1: Fetching List Data ===\n")
    all_items = []

    for date_str in fetch_dates:
        print(f"--- Date: {date_str} ---")
        date_items = []

        for cat_name, cat_code in CATEGORY_CODES.items():
            page = 1
            cat_items = []

            while True:
                payload = {
                    "pageNo": page, "pageSize": 100,
                    "categoryCode": cat_code, "isGov": True,
                    "excludeDistrictPrefix": ["90", "006011", "H0", "001111"],
                    "publishDateBegin": date_str, "publishDateEnd": date_str,
                    "_t": int(datetime.now().timestamp() * 1000)
                }
                body = json.dumps(payload).encode('utf-8')
                req = urllib.request.Request(LIST_API, data=body, headers=API_HEADERS)
                try:
                    with urllib.request.urlopen(req, timeout=30) as resp:
                        result = json.loads(resp.read().decode('utf-8'))
                except Exception as e:
                    print(f"  [{cat_name}] Request failed: {e}")
                    break

                if not result.get('success'):
                    break
                data = result.get('result', {}).get('data', {})
                if not data:
                    break
                total = data.get('total', 0)
                items = data.get('data', [])

                for item in items:
                    dc = item.get('districtCode', '')
                    dn = item.get('districtName', '')
                    if dc in TARGET_DISTRICT_CODES or any(d in dn for d in TARGET_DISTRICTS):
                        cat_items.append(item)

                if len(items) < 100 or page * 100 >= total:
                    break
                page += 1
                time.sleep(0.2)

            if cat_items:
                print(f"  [{cat_name}] Target district: {len(cat_items)} records")
                date_items.extend(cat_items)
            time.sleep(0.2)

        # Deduplicate by articleId
        seen = set()
        unique_items = []
        for item in date_items:
            aid = item.get('articleId', '')
            if aid and aid not in seen:
                seen.add(aid)
                unique_items.append(item)
        print(f"  Total after dedup: {len(unique_items)}\n")
        all_items.extend(unique_items)

    all_items.sort(key=lambda x: x.get('publishDate', 0) or 0, reverse=True)
    print(f"Grand total: {len(all_items)} records\n")
    return all_items


# ========================================================================
# Stage 2: Detail Extraction
# ========================================================================

def fetch_details(all_items):
    """Fetch detail content and extract structured fields for each item."""
    print("=== Stage 2: Fetching Details ===\n")

    for idx, item in enumerate(all_items):
        aid = item.get('articleId', '')
        if not aid:
            continue
        pn = (item.get('projectName') or '')[:40]
        print(f"  [{idx+1}/{len(all_items)}] {pn}...")

        detail_url = (
            f'{DETAIL_API}?articleId={urllib.parse.quote(aid)}'
            f'&timestamp={int(datetime.now().timestamp()*1000)}'
        )
        try:
            req = urllib.request.Request(detail_url)
            with urllib.request.urlopen(req, timeout=30) as resp:
                detail = json.loads(resp.read().decode('utf-8'))
        except Exception as e:
            print(f"    Request failed: {e}")
            continue

        if not detail.get('success') or not detail.get('result', {}).get('data'):
            continue

        d = detail['result']['data']
        content = d.get('content', '')
        item['_raw_content'] = content
        text = strip_html(content)
        item['_announcement_links'] = d.get('announcementLinkDtoList', [])

        cat_type = PATH_TO_CATEGORY.get(item.get('pathName', ''), '')

        # Void/废标 detection
        norm_text = re.sub(r'\s+', ' ', text[:3000])
        if re.search(r'废标公告|招标失败公告|采购失败公告|流标公告', norm_text):
            item['_override_type'] = '废标'
        elif re.search(r'废标|招标失败|采购失败|终止采购|取消采购|流标', norm_text):
            has_winning = bool(re.search(
                r'中标[（(]成交[)）]信息|中标[（(]成交[)）]结果|中标供应商名称', norm_text
            ))
            has_failure_only = bool(re.search(r'废标理由|废标结果|流标原因', norm_text)) and not has_winning
            if has_failure_only:
                item['_override_type'] = '废标'

        cat_type = item.get('_override_type', cat_type)

        # Extract fields
        if cat_type != '采购合同公告':
            budget = extract_budget(text)
            item['_budget'] = budget if budget and budget > 100 else None
        win_amount = extract_winning_amount(text)
        item['_winning_amount'] = win_amount if win_amount and win_amount > 0 else None

        # Supplier should only be extracted for post-award stages
        if cat_type not in ('采购项目公告', '采购意向', '意见征询'):
            supplier = extract_supplier(text)
            item['_supplier'] = supplier or item.get('supplierName')
        else:
            item['_supplier'] = None

        agency = extract_agency(text, cat_type)
        if agency:
            item['_agency'] = agency
        elif cat_type in ('采购意向', '意见征询'):
            # Agency not yet determined at intention/consultation stage
            item['_agency'] = ''
        elif cat_type in ('采购合同公告', '电子卖场公告'):
            item['_agency'] = ''
        else:
            # 采购项目公告, 采购结果公告, 废标, etc. -> derive from author
            author = item.get('author', '')
            item['_agency'] = author if any(kw in author for kw in ['公司', '事务所', '代理']) else ''
        purchaser = extract_purchaser(text)
        item['_purchaser'] = purchaser or item.get('purchaseName', '')
        time.sleep(0.3)


# ========================================================================
# Stage 2.5: Cross-Referencing
# ========================================================================

def same_project_inherit(all_items):
    """Layer 0: Inherit fields from earlier announcements of the same project within this batch."""
    # Types that should NEVER receive supplier (pre-award stages)
    NO_SUPPLIER_TYPES = ('采购项目公告', '采购意向', '意见征询')
    # Types that should NEVER receive agency (not yet determined at this stage)
    NO_AGENCY_TYPES = ('采购意向', '意见征询')
    
    # Group by projectName
    by_name = defaultdict(list)
    for idx, item in enumerate(all_items):
        name = (item.get('projectName') or item.get('title') or '').strip()
        if name:
            by_name[name].append((idx, item))

    inherit_count = {'budget': 0, 'agency': 0, 'supplier': 0, 'district': 0, 'purchaser': 0}

    for name, indexed in by_name.items():
        if len(indexed) < 2:
            continue

        # Sort by publishDate ASC (earliest first)
        indexed.sort(key=lambda x: x[1].get('publishDate', 0))

        # Walk in time order: accumulate source values, reset on 废标
        src_budget = src_agency = src_supplier = src_district = src_purchaser = None
        for idx, item in indexed:
            cat_type = item.get('_override_type') or PATH_TO_CATEGORY.get(item.get('pathName', ''), '')

            # 废标 breaks the inheritance chain — reset accumulated source
            if cat_type == '废标':
                src_budget = src_agency = src_supplier = src_district = src_purchaser = None
                continue

            # Inherit missing fields from accumulated source (earlier items only)
            if not item.get('_budget') or item['_budget'] <= 100:
                if src_budget:
                    item['_budget'] = src_budget
                    inherit_count['budget'] += 1
            if not item.get('_agency'):
                if src_agency and cat_type not in NO_AGENCY_TYPES:
                    item['_agency'] = src_agency
                    inherit_count['agency'] += 1
            if not item.get('_supplier'):
                if src_supplier and cat_type not in NO_SUPPLIER_TYPES:
                    item['_supplier'] = src_supplier
                    inherit_count['supplier'] += 1
            if not item.get('districtCode'):
                if src_district:
                    item['districtCode'] = src_district
                    inherit_count['district'] += 1
            if not item.get('_purchaser') and not item.get('purchaseName'):
                if src_purchaser:
                    item['_purchaser'] = src_purchaser
                    inherit_count['purchaser'] += 1

            # Accumulate source values from this item for later items
            # Budget should only come from the procurement announcement (采购项目公告);
            # result/contract amounts are not the procurement budget.
            if src_budget is None and item.get('_budget') and item['_budget'] > 100 and cat_type == '采购项目公告':
                src_budget = item['_budget']
            if src_agency is None and item.get('_agency'):
                src_agency = item['_agency']
            if src_supplier is None and item.get('_supplier'):
                src_supplier = item['_supplier']
            if src_district is None and item.get('districtCode'):
                src_district = item['districtCode']
            if src_purchaser is None and (item.get('_purchaser') or item.get('purchaseName')):
                src_purchaser = item.get('_purchaser') or item.get('purchaseName')

    if any(inherit_count.values()):
        print(f"  Same-project inherit: {inherit_count}")
    return


def inherit_from_main(all_items, main_file_path):
    """Layer 0b: Inherit fields from the main Excel file's historical records.
    
    Rules:
    1. Walk project history in time order, accumulate source values
    2. 废标 resets the accumulation chain
    3. Current batch items inherit from accumulated history (past→present)
    """
    try:
        wb = openpyxl.load_workbook(main_file_path)  # read_only=True removed due to Python 3.13 crash
        ws = wb.active

        # Read all main file rows and group by project name
        from collections import defaultdict as dd
        main_groups = dd(list)
        for r in range(4, ws.max_row + 1):
            name = str(ws.cell(r, 5).value or '').strip()
            if not name:
                continue
            main_groups[name].append({
                'pub_time': ws.cell(r, 1).value,
                'pub_type': ws.cell(r, 2).value,
                'budget': ws.cell(r, 6).value,
                'agency': ws.cell(r, 9).value,
                'supplier': ws.cell(r, 8).value,
                'purchaser': ws.cell(r, 4).value,
                'district': ws.cell(r, 3).value,
            })
        wb.close()

        # For each project, walk in time order and accumulate source (skip 废标)
        history = {}
        for name, entries in main_groups.items():
            entries.sort(key=lambda x: x['pub_time'] if x['pub_time'] else '')
            src = {'budget': None, 'agency': None, 'supplier': None, 'purchaser': None, 'district': None}
            for e in entries:
                if e['pub_type'] == '废标':
                    src = {'budget': None, 'agency': None, 'supplier': None, 'purchaser': None, 'district': None}
                    continue
                # Budget should only come from the procurement announcement (采购项目公告)
                if src['budget'] is None and isinstance(e['budget'], (int, float)) and e['budget'] > 100 and e['pub_type'] == '采购项目公告':
                    src['budget'] = e['budget']
                if src['agency'] is None and e['agency']:
                    src['agency'] = e['agency']
                if src['supplier'] is None and e['supplier']:
                    src['supplier'] = e['supplier']
                if src['purchaser'] is None and e['purchaser']:
                    src['purchaser'] = e['purchaser']
                if src['district'] is None and e['district']:
                    src['district'] = e['district']
            if any(v is not None for v in src.values()):
                history[name] = src

        count = {'budget': 0, 'agency': 0, 'supplier': 0, 'district': 0, 'purchaser': 0}
        for item in all_items:
            name = (item.get('projectName') or item.get('title') or '').strip()
            if not name or name not in history:
                continue
            h = history[name]

            cat_type = item.get('_override_type') or PATH_TO_CATEGORY.get(item.get('pathName', ''), '')
            if not item.get('_budget') or item['_budget'] <= 100:
                if h['budget']:
                    item['_budget'] = h['budget']
                    count['budget'] += 1
            if not item.get('_agency'):
                if h['agency'] and cat_type not in ('采购意向', '意见征询'):
                    item['_agency'] = h['agency']
                    count['agency'] += 1
            if not item.get('_supplier'):
                if h['supplier'] and cat_type not in ('采购项目公告', '采购意向', '意见征询'):
                    item['_supplier'] = h['supplier']
                    count['supplier'] += 1
            if not item.get('districtCode'):
                if h['district']:
                    rev = {v: k for k, v in DISTRICT_MAP.items()}
                    code = rev.get(h['district']) or rev.get('浙江省' + h['district']) or rev.get(CITY_NAME + h['district'])
                    if code:
                        item['districtCode'] = code
                        count['district'] += 1
            if not item.get('_purchaser') and not item.get('purchaseName'):
                if h['purchaser']:
                    item['_purchaser'] = h['purchaser']
                    count['purchaser'] += 1

        if any(count.values()):
            print(f"  Main-file inherit: {count}")
    except Exception as e:
        print(f"  Main-file inherit skipped: {e}")
    return


def cross_reference(all_items, main_file_path=None):
    """Cross-reference missing data from linked announcements + main file history."""
    print("\n=== Stage 2.5: Cross-Referencing ===\n")

    # Layer 0: Same-project inheritance within current batch
    same_project_inherit(all_items)

    # Layer 0b: Inherit from main file historical data
    if main_file_path and os.path.exists(main_file_path):
        inherit_from_main(all_items, main_file_path)

    print(f"  Starting individual cross-ref for {len(all_items)} items...")
    cr_count = 0
    # Limit individual cross-ref to avoid API rate issues with large batches
    max_cross_ref = 50
    if len(all_items) > max_cross_ref:
        print(f"  Skipping individual cross-ref ({len(all_items)} items > {max_cross_ref} limit, using main-file inherit only)")
        return
    for idx, item in enumerate(all_items):
        cat_type = item.get('_override_type') or PATH_TO_CATEGORY.get(item.get('pathName', ''), '')
        if cat_type == '废标':
            continue

        # Layer 1: Announcement link chain
        links = item.get('_announcement_links') or []
        for link in links:
            linked_id = link.get('articleId', '')
            if not linked_id or linked_id == item.get('articleId'):
                continue
            try:
                durl = f'{DETAIL_API}?articleId={urllib.parse.quote(linked_id)}&timestamp={int(datetime.now().timestamp()*1000)}'
                req = urllib.request.Request(durl)
                with urllib.request.urlopen(req, timeout=30) as resp:
                    d2 = json.loads(resp.read().decode('utf-8'))
                if d2.get('success') and d2.get('result', {}).get('data'):
                    dd = d2['result']['data']
                    t2 = strip_html(dd.get('content', ''))

                    if item.get('_budget') is None or item['_budget'] <= 10000:
                        b2 = extract_budget(t2)
                        if b2 and b2 > 100 and (item.get('_budget') is None or b2 > item['_budget']):
                            item['_budget'] = b2
                            print(f"  Cross-ref: {(item.get('projectName') or '')[:30]}... -> budget {b2:.0f}")

                    if not item.get('_supplier'):
                        s2 = extract_supplier(t2)
                        if s2:
                            item['_supplier'] = s2
                            print(f"  Cross-ref: {(item.get('projectName') or '')[:30]}... -> supplier {s2}")

                    if not item.get('_agency'):
                        a2 = extract_agency(t2, '采购项目公告')
                        if not a2:
                            auth = dd.get('author', '')
                            a2 = auth if any(kw in auth for kw in ['公司', '事务所', '代理']) else None
                        if a2:
                            item['_agency'] = a2
                            print(f"  Cross-ref: {(item.get('projectName') or '')[:30]}... -> agency {a2}")
            except Exception:
                pass
            time.sleep(0.3)

        # Layer 2: Result fallback (use winning amount as budget)
        if cat_type == '采购结果公告' and (item.get('_budget') is None or item.get('_budget', 0) <= 10000):
            wa = item.get('_winning_amount')
            if wa and wa > 100:
                item['_budget'] = wa
                print(f"  Fallback: {(item.get('projectName') or '')[:30]}... -> budget {wa:.0f}")

        # Layer 3: URL extraction from HTML
        if item.get('_budget') is None or item.get('_budget', 0) <= 10000:
            raw = item.get('_raw_content', '')
            linked_ids = re.findall(r'articleId=([A-Za-z0-9+/=]+(?:%3D%3D|%3D))', raw)
            linked_ids += re.findall(r'articleId=([A-Za-z0-9+/]+=*)', urllib.parse.unquote(raw))
            seen = set()
            for raw_id in linked_ids:
                did = urllib.parse.unquote(raw_id)
                if did in seen or did == item.get('articleId'):
                    continue
                seen.add(did)
                try:
                    durl = f'{DETAIL_API}?articleId={urllib.parse.quote(did)}&timestamp={int(datetime.now().timestamp()*1000)}'
                    req = urllib.request.Request(durl)
                    with urllib.request.urlopen(req, timeout=30) as resp:
                        d3 = json.loads(resp.read().decode('utf-8'))
                    if d3.get('success') and d3.get('result', {}).get('data'):
                        t3 = strip_html(d3['result']['data'].get('content', ''))
                        b3 = extract_budget(t3)
                        if b3 and b3 > 100 and (item.get('_budget') is None or b3 > item['_budget']):
                            item['_budget'] = b3
                            print(f"  URL-ref: {(item.get('projectName') or '')[:30]}... -> budget {b3:.0f}")
                            break
                        if not item.get('_agency'):
                            a3 = extract_agency(t3, '采购项目公告')
                            if not a3:
                                auth = d3['result']['data'].get('author', '')
                                a3 = auth if any(kw in auth for kw in ['公司', '事务所', '代理']) else None
                            if a3:
                                item['_agency'] = a3
                except Exception:
                    pass
                time.sleep(0.3)
        if (idx + 1) % 20 == 0:
            print(f"  Cross-ref progress: {idx+1}/{len(all_items)}")

    # Contract fallback
    for item in all_items:
        cat_type = item.get('_override_type') or PATH_TO_CATEGORY.get(item.get('pathName', ''), '')
        if cat_type == '采购合同公告' and (item.get('_budget') is None or item['_budget'] < 100):
            wa = item.get('_winning_amount')
            if wa and wa > 100:
                item['_budget'] = wa
                print(f"  Contract: {(item.get('projectName') or '')[:30]}... -> budget {wa:.0f}")


# ========================================================================
# Stage 3: Excel Generation
# ========================================================================

def generate_excel(all_items, fetch_dates):
    """Generate the output Excel file."""
    print(f"\n=== Stage 3: Generating Excel ===\n")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '采购公告'

    headers = [
        '序号', '公告时间', '公告类型', '所在区县', '采购人', '采购项目名称',
        '预算金额(元)', '中标金额(元)', '中标单位', '代理机构', '地址链接URL'
    ]

    # Filename（按地市命名：{地市}_采购公告_{日期}.xlsx，落盘到 workspace，
    # 与 merge_to_main 在同一采集工作区自动衔接，多城市同目录不冲突）
    if len(fetch_dates) == 1:
        date_part = fetch_dates[0]
    else:
        date_part = f"{min(fetch_dates)}_{max(fetch_dates)}"
    out_file = str(Path(WORKSPACE_DIR) / f'{CITY_BASE}_采购公告_{date_part}.xlsx')

    # Header
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    # Sort by publish time descending (newest first) for the output sheet
    all_items.sort(key=lambda x: x.get('publishDate', 0) or 0, reverse=True)

    for idx, item in enumerate(all_items, 1):
        pub_ts = item.get('publishDate', 0)
        if pub_ts:
            dt = datetime.fromtimestamp(pub_ts / 1000)
            pub_time = dt.strftime('%Y-%m-%d %H:%M')
        else:
            pub_time = ''

        cat_type = item.get('_override_type') or PATH_TO_CATEGORY.get(item.get('pathName', ''), '')

        dc = item.get('districtCode', '')
        district = DISTRICT_MAP.get(dc, item.get('districtName', ''))
        # 区县名按地市动态剥离"浙江省{地市}"前缀（原硬编码"浙江省丽水市"泛化）
        if district.startswith(f'浙江省{CITY_NAME}'):
            district = district.replace(f'浙江省{CITY_NAME}', '')

        aid = item.get('articleId', '')
        link = f'https://zfcg.czt.zj.gov.cn/site/detail?articleId={urllib.parse.quote(aid)}' if aid else ''

        row = [
            idx, pub_time, cat_type, district,
            item.get('_purchaser', item.get('purchaseName', '')),
            item.get('projectName', item.get('title', '')),
            item.get('_budget'),
            item.get('_winning_amount'),
            item.get('_supplier', ''),
            item.get('_agency', ''),
            link,
        ]
        for c, val in enumerate(row, 1):
            if c == 11 and link:
                cell = ws.cell(idx + 1, c)
                cell.value = link
                cell.hyperlink = link
                cell.style = 'Hyperlink'
            else:
                ws.cell(idx + 1, c, val)

    # Summary row
    summary_row = len(all_items) + 2
    ws.cell(summary_row, 1,
            f'共 {len(all_items)} 条公告 | 采集时间：{datetime.now().strftime("%Y-%m-%d")} | '
            f'数据来源：浙江政府采购网')

    # Column widths
    col_widths = {'A': 6, 'B': 16, 'C': 14, 'D': 14, 'E': 25,
                  'F': 45, 'G': 14, 'H': 14, 'I': 25, 'J': 25, 'K': 50}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    wb.save(out_file)
    print(f"  Saved: {out_file}")
    print(f"  Total records: {len(all_items)}")

    # Stats
    print(f"\n=== Correction Statistics ===")
    stats = {'budget': 0, 'amount': 0, 'agency': 0, 'supplier': 0, 'type': 0}
    for item in all_items:
        if item.get('_budget') and item.get('_budget') > 100:
            bp = item.get('budgetPrice')
            if not bp or bp == '0' or bp == 0:
                stats['budget'] += 1
        if item.get('_winning_amount') and item.get('_winning_amount') > 0:
            if not item.get('totalContractAmount'):
                stats['amount'] += 1
        if item.get('_agency'):
            author = item.get('author', '')
            if item['_agency'] != author:
                stats['agency'] += 1
        if item.get('_supplier') and item['_supplier'] != item.get('supplierName'):
            stats['supplier'] += 1
        if item.get('_override_type'):
            stats['type'] += 1

    for k, v in stats.items():
        if v > 0:
            print(f"  {k}: {v} corrected")
    return out_file


# ========================================================================
# Main
# ========================================================================

def main():
    parser = argparse.ArgumentParser(
        description='政府采购公告数据采集脚本 (ZFCG Data Collector)',
        epilog='Example: python fetch_zfcg.py --date 2026-06-22 --workspace C:\\path\\to\\workspace'
    )
    parser.add_argument('--workspace', type=str, default='.',
                        help='工作目录（主文件所在目录），默认当前目录 .')
    parser.add_argument('--city', type=str, default='丽水市',
                        help='目标地市（默认丽水市），须在 references/districts.json 中')
    parser.add_argument('--date', type=str, help='Target date (YYYY-MM-DD)')
    parser.add_argument('--range', nargs=2, type=str, metavar=('START', 'END'),
                        help='Date range (YYYY-MM-DD YYYY-MM-DD)')
    parser.add_argument('--verify', action='store_true', help='Run validation after collection')
    args = parser.parse_args()

    # 按 --city 加载地市区县映射并动态生成目标集合（默认丽水市，向后兼容）
    global MAIN_FILE, WORKSPACE_DIR, CITY_NAME, CITY_BASE, DISTRICT_MAP, TARGET_DISTRICTS, TARGET_DISTRICT_CODES
    CITY_NAME = args.city
    CITY_BASE = CITY_NAME[:-1] if CITY_NAME.endswith('市') else CITY_NAME
    DISTRICT_MAP = load_districts(CITY_NAME)
    TARGET_DISTRICTS = set(DISTRICT_MAP.values())
    TARGET_DISTRICT_CODES = set(DISTRICT_MAP.keys())

    # Set up paths from --workspace
    workspace = Path(args.workspace).resolve()
    WORKSPACE_DIR = str(workspace)
    MAIN_FILE = str(workspace / f'{CITY_BASE}政府采购公告.xlsx')
    if not workspace.exists():
        sys.exit(f'错误：工作目录不存在：{workspace}\n请用 --workspace 指定包含主文件的目录')

    # Check main file lock (Excel opened)
    lock_file = workspace / f'~${CITY_BASE}政府采购公告.xlsx'
    if lock_file.exists():
        sys.exit(f'错误：主文件被 Excel 打开，无法读取！\n锁文件：{lock_file}\n请先关闭 Excel 再运行。')

    print("=" * 70)
    print("政府采购公告数据采集 (ZFCG Data Collector)")
    print("=" * 70)
    print()

    # Determine dates
    if args.date:
        fetch_dates = [args.date]
        print(f"Manual date: {args.date}\n")
    elif args.range:
        start = datetime.strptime(args.range[0], '%Y-%m-%d')
        end = datetime.strptime(args.range[1], '%Y-%m-%d')
        fetch_dates = []
        d = start
        while d <= end:
            fetch_dates.append(d.strftime('%Y-%m-%d'))
            d += timedelta(days=1)
        print(f"Manual range: {args.range[0]} ~ {args.range[1]} ({len(fetch_dates)} days)\n")
    else:
        fetch_dates = analyze_main_file()

    if not fetch_dates:
        print("No dates to fetch. Exiting.")
        return

    # Stage 1: Fetch list
    all_items = fetch_list_data(fetch_dates)
    if not all_items:
        print("No announcements found. Exiting.")
        return

    # Stage 2: Fetch details
    fetch_details(all_items)

    # Stage 2.5: Cross-reference
    cross_reference(all_items, MAIN_FILE)

    # Stage 3: Generate Excel
    out_file = generate_excel(all_items, fetch_dates)

    print(f"\n{'='*70}")
    print("Collection complete!")
    print(f"  Output:  {out_file}")
    print(f"  Records: {len(all_items)}")
    print(f"  Dates:   {', '.join(fetch_dates)}")
    print(f"{'='*70}")


if __name__ == '__main__':
    main()
