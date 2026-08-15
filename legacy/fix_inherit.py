# -*- coding: utf-8 -*-
"""
一次性修复主表中同项目可继承的缺失字段。
规则：时间顺序单向继承 + 废标断链 + pre-award 不继承中标单位/代理机构

用法：
  python fix_inherit.py --workspace C:\\path\\to\\workspace
"""
import openpyxl
import sys
import argparse
from collections import defaultdict
from pathlib import Path

NO_SUPPLIER_TYPES = ('采购项目公告', '采购意向', '意见征询')
NO_AGENCY_TYPES = ('采购意向', '意见征询')


def get_args():
    parser = argparse.ArgumentParser(
        description='修复主表中同项目可继承的缺失字段',
        epilog='Example: python fix_inherit.py --workspace C:\\path\\to\\workspace'
    )
    parser.add_argument('--workspace', type=str, default='.',
                        help='工作目录（主文件所在），默认当前目录 .')
    return parser.parse_args()


def main():
    args = get_args()
    workspace = Path(args.workspace).resolve()
    main_file = workspace / '丽水政府采购公告.xlsx'

    if not workspace.exists():
        sys.exit(f'错误：工作目录不存在：{workspace}')
    if not main_file.exists():
        sys.exit(f'错误：主文件不存在：{main_file}')

    wb = openpyxl.load_workbook(main_file)
    ws = wb.active

    # 读取全部数据
    rows = []
    for r in range(4, ws.max_row + 1):
        rows.append({
            'row': r,
            'pub_time': ws.cell(r, 1).value,
            'pub_type': ws.cell(r, 2).value,
            'district': ws.cell(r, 3).value,
            'purchaser': ws.cell(r, 4).value,
            'project_name': str(ws.cell(r, 5).value or '').strip(),
            'budget': ws.cell(r, 6).value,
            'winning_amount': ws.cell(r, 7).value,
            'supplier': ws.cell(r, 8).value,
            'agency': ws.cell(r, 9).value,
            'url': ws.cell(r, 10).value,
        })

    # 按项目名分组
    groups = defaultdict(list)
    for entry in rows:
        name = entry['project_name']
        if name:
            groups[name].append(entry)

    stats = {'budget': 0, 'agency': 0, 'supplier': 0, 'purchaser': 0, 'district': 0}

    for name, items in groups.items():
        if len(items) < 2:
            continue

        # Sort by pub_time ASC (earliest first)
        items.sort(key=lambda x: x['pub_time'] if x['pub_time'] else '')

        # Walk in time order: accumulate source values, reset on 废标
        src_budget = src_agency = src_supplier = src_purchaser = src_district = None
        for item in items:
            # 废标 breaks the inheritance chain
            if item['pub_type'] == '废标':
                src_budget = src_agency = src_supplier = src_purchaser = src_district = None
                continue

            # Inherit missing fields from accumulated source (earlier items only)
            missing_budget = not isinstance(item['budget'], (int, float)) or item['budget'] <= 100
            if missing_budget and src_budget:
                ws.cell(item['row'], 6, src_budget)
                stats['budget'] += 1
            if not item['agency'] and src_agency and item['pub_type'] not in NO_AGENCY_TYPES:
                ws.cell(item['row'], 9, src_agency)
                stats['agency'] += 1
            if not item['supplier'] and src_supplier and item['pub_type'] not in NO_SUPPLIER_TYPES:
                ws.cell(item['row'], 8, src_supplier)
                stats['supplier'] += 1
            if not item['purchaser'] and src_purchaser:
                ws.cell(item['row'], 4, src_purchaser)
                stats['purchaser'] += 1
            if not item['district'] and src_district:
                ws.cell(item['row'], 3, src_district)
                stats['district'] += 1

            # Accumulate source values from this item for later items
            if src_budget is None and isinstance(item['budget'], (int, float)) and item['budget'] > 100:
                src_budget = item['budget']
            if src_agency is None and item['agency']:
                src_agency = item['agency']
            if src_supplier is None and item['supplier']:
                src_supplier = item['supplier']
            if src_purchaser is None and item['purchaser']:
                src_purchaser = item['purchaser']
            if src_district is None and item['district']:
                src_district = item['district']

    wb.save(main_file)
    wb.close()
    print(f'修复完成: {stats}')


if __name__ == '__main__':
    main()
