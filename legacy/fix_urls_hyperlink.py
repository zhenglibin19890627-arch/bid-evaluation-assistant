# -*- coding: utf-8 -*-
"""
一次性将主文件中所有URL文本转换为可点击超链接

用法：
  python fix_urls_hyperlink.py --workspace C:\\path\\to\\workspace
"""
import openpyxl
import sys
import argparse
from pathlib import Path


def get_args():
    parser = argparse.ArgumentParser(
        description='主文件 URL 文本 → 超链接',
        epilog='Example: python fix_urls_hyperlink.py --workspace C:\\path\\to\\workspace'
    )
    parser.add_argument('--workspace', type=str, default='.',
                        help='工作目录（主文件所在），默认当前目录 .')
    return parser.parse_args()


def main():
    args = get_args()
    workspace = Path(args.workspace).resolve()
    main_file = workspace / '丽水政府采购公告.xlsx'

    if not workspace.exists():
        sys.exit(f'错误：工作目录不存在：{workspace}')
    if not main_file.exists():
        sys.exit(f'错误：主文件不存在：{main_file}')

    wb = openpyxl.load_workbook(main_file)
    ws = wb.active

    count = 0
    for r in range(4, ws.max_row + 1):
        url = ws.cell(r, 10).value
        if url and isinstance(url, str) and url.startswith('http'):
            cell = ws.cell(r, 10)
            cell.hyperlink = url
            cell.value = url
            cell.style = 'Hyperlink'
            count += 1

    wb.save(main_file)
    wb.close()
    print(f'已转换 {count} 个URL为超链接')


if __name__ == '__main__':
    main()
