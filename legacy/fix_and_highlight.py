# -*- coding: utf-8 -*-
"""
修复两条特定历史记录 + 对数据缺失单元格高亮（蓝色）
输出新文件 `丽水政府采购公告_高亮版.xlsx`，不覆盖主文件

用法：
  python fix_and_highlight.py --workspace C:\\path\\to\\workspace
"""
import openpyxl
import sys
import argparse
from openpyxl.styles import PatternFill
from datetime import datetime
from pathlib import Path


def get_args():
    parser = argparse.ArgumentParser(
        description='修复特定记录 + 高亮数据缺失',
        epilog='Example: python fix_and_highlight.py --workspace C:\\path\\to\\workspace'
    )
    parser.add_argument('--workspace', type=str, default='.',
                        help='工作目录（主文件所在），默认当前目录 .')
    return parser.parse_args()


def main():
    args = get_args()
    workspace = Path(args.workspace).resolve()
    main_file = workspace / '丽水政府采购公告.xlsx'

    if not workspace.exists():
        sys.exit(f'错误：工作目录不存在：{workspace}')
    if not main_file.exists():
        sys.exit(f'错误：主文件不存在：{main_file}')

    BLUE_FILL = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')

    wb = openpyxl.load_workbook(main_file)
    ws = wb.active

    # ── Step 1: Fix two specific records ──

    # Scan for the two URLs and apply fixes
    for r in range(4, ws.max_row + 1):
        url = str(ws.cell(r, 10).value or '')

        # URL#1: xgcHm9QWEM0kh
        if 'xgcHm9QWEM0kh' in url:
            ws.cell(r, 8).value = '丽水市启点印业有限公司'  # supplier
            print(f'Row {r}: supplier → 丽水市启点印业有限公司')

        # URL#2: FHcPWmD1mgpiDjI6LNx
        if 'FHcPWmD1mgpiDjI6LNx' in url:
            ws.cell(r, 7).value = 245000  # win_amt
            ws.cell(r, 8).value = '浙江联和安保集团有限公司'  # supplier
            print(f'Row {r}: win_amt → 245000, supplier → 浙江联和安保集团有限公司')

    # ── Step 2: Highlight data quality issues in blue ──

    highlight_count = 0
    cell_count = 0

    for r in range(4, ws.max_row + 1):
        typ = str(ws.cell(r, 2).value or '')
        supplier = ws.cell(r, 8).value
        win_amt = ws.cell(r, 7).value
        budget = ws.cell(r, 6).value
        agency = ws.cell(r, 9).value
        buyer = ws.cell(r, 4).value

        highlighted_this_row = False

        # Supplier missing → mark col 8 (中标单位)
        if typ in ['采购结果公告', '采购合同公告']:
            if supplier is None or supplier == '':
                ws.cell(r, 8).fill = BLUE_FILL
                cell_count += 1
                highlighted_this_row = True

        # Win amount missing → mark col 7 (中标金额)
        if typ in ['采购结果公告', '采购合同公告']:
            if win_amt is None or win_amt == '' or win_amt == 0:
                ws.cell(r, 7).fill = BLUE_FILL
                cell_count += 1
                highlighted_this_row = True

        # Budget missing → mark col 6 (预算金额)
        if budget is None or budget == '' or budget == 0:
            ws.cell(r, 6).fill = BLUE_FILL
            cell_count += 1
            highlighted_this_row = True

        # Agency missing for project/correction types → mark col 9
        if typ in ['采购项目公告', '更正公告']:
            if agency is None or agency == '':
                ws.cell(r, 9).fill = BLUE_FILL
                cell_count += 1
                highlighted_this_row = True

        # Buyer missing → mark col 4 (采购人)
        if buyer is None or buyer == '':
            ws.cell(r, 4).fill = BLUE_FILL
            cell_count += 1
            highlighted_this_row = True

        if highlighted_this_row:
            highlight_count += 1

    # Save to a separate file
    out_file = workspace / '丽水政府采购公告_高亮版.xlsx'
    wb.save(out_file)
    wb.close()
    print(f'Saved to: {out_file.name}')

    print(f'\nDone!')
    print(f'  Fixed: 2 records (supplier + win_amt)')
    print(f'  Highlighted: {highlight_count} rows with {cell_count} blue cells')


if __name__ == '__main__':
    main()
