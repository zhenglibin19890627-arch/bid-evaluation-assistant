# -*- coding: utf-8 -*-
"""
一键修复主文件：字符串时间 → 文本 'YYYY-MM-DD HH:MM' + 降序排序

融合适配（方案 8）：排序重写改为搬运 1–12 列（10 数据列 + 初筛状态/研判状态），
状态列随行走；--city 参数（默认"丽水市"），主文件按地市命名。

用法：
  python fix_format.py --workspace C:\\path\\to\\workspace
  python fix_format.py --city 杭州市 --workspace C:\\path\\to\\workspace
"""
import openpyxl
import sys
import argparse
from datetime import datetime
from pathlib import Path

HEADER_ROW = 3
FIRST_DATA_ROW = 4

# 主文件 12 列：10 数据列 + 初筛状态/研判状态
ALL_COLUMNS = [
    '公告时间', '公告类型', '所在区县', '采购人', '采购项目名称',
    '预算金额(元)', '中标金额(元)', '中标单位', '代理机构', '地址链接URL',
    '初筛状态', '研判状态',
]


def get_args():
    parser = argparse.ArgumentParser(
        description='主文件时间格式归一 + 降序排序',
        epilog='Example: python fix_format.py --workspace C:\\path\\to\\workspace'
    )
    parser.add_argument('--workspace', type=str, default='.',
                        help='工作目录（主文件所在），默认当前目录 .')
    parser.add_argument('--city', type=str, default='丽水市',
                        help='目标地市（默认丽水市），主文件按地市命名')
    return parser.parse_args()


def main():
    args = get_args()
    workspace = Path(args.workspace).resolve()
    city = args.city
    city_base = city[:-1] if city.endswith('市') else city
    main_file = workspace / f'{city_base}政府采购公告.xlsx'

    if not workspace.exists():
        sys.exit(f'错误：工作目录不存在：{workspace}')
    if not main_file.exists():
        sys.exit(f'错误：主文件不存在：{main_file}')

    wb = openpyxl.load_workbook(main_file)
    ws = wb.active

    # 读取所有数据行（1–12 列，含状态列）
    headers = [ws.cell(HEADER_ROW, c).value for c in range(1, len(ALL_COLUMNS) + 1)]
    print(f'Headers: {headers}')

    data_rows = []
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, len(ALL_COLUMNS) + 1)]
        if vals[0]:  # has time
            data_rows.append(vals)

    # 转换字符串时间为 datetime
    fixed = 0
    for row in data_rows:
        if isinstance(row[0], str) and len(row[0]) >= 16:
            try:
                row[0] = datetime.strptime(row[0][:16], '%Y-%m-%d %H:%M')
                fixed += 1
            except ValueError:
                pass

    print(f'Converted {fixed} string times to datetime')
    print(f'Total data rows: {len(data_rows)}')

    # 排序（最新在前）
    def _sort_key(val):
        if isinstance(val, datetime):
            return val
        if isinstance(val, str) and len(val) >= 16:
            try:
                return datetime.strptime(val[:16], '%Y-%m-%d %H:%M')
            except ValueError:
                pass
        return datetime.min
    data_rows.sort(key=lambda r: _sort_key(r[0]), reverse=True)

    # 清空并写回（搬运 1–12 列，状态列随行走）
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        for c in range(1, len(ALL_COLUMNS) + 1):
            ws.cell(r, c).value = None

    for i, row in enumerate(data_rows):
        r = FIRST_DATA_ROW + i
        for c, val in enumerate(row, 1):
            if c == 1 and isinstance(val, datetime):
                val = val.strftime('%Y-%m-%d %H:%M')
            ws.cell(r, c, val)
        # 时间列格式（@ 表示文本）
        ws.cell(r, 1).number_format = '@'

    # 更新时间戳
    ws.cell(2, 1).value = f'采集时间：{datetime.now().strftime("%Y-%m-%d %H:%M")} | 数据来源：浙江政府采购网 | 共 {len(data_rows)} 条'

    wb.save(main_file)
    wb.close()

    print('Done. Main file sorted and normalized.')


if __name__ == '__main__':
    main()
